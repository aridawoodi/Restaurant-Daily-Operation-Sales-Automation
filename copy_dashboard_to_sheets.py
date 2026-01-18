"""
Copy a local Excel sheet (with formulas and formatting) to a Google Sheet tab.
Prompts for source sheet name and destination sheet name.
"""

import os
from datetime import date, datetime
from typing import Any, Dict, Optional, Tuple

from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import load_workbook
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils import column_index_from_string
from openpyxl.utils.datetime import to_excel

from csv_to_sheets import CSVToSheetsAutomation


BORDER_STYLE_MAP = {
    "thin": "SOLID",
    "medium": "SOLID_MEDIUM",
    "thick": "SOLID_THICK",
    "double": "DOUBLE",
    "dashed": "DASHED",
    "dotted": "DOTTED",
    "hair": "DOTTED",
    "mediumDashed": "DASHED",
    "dashDot": "DASHED",
    "dashDotDot": "DASHED",
    "slantDashDot": "DASHED",
}


def prompt_with_default(prompt: str, default: str, allow_blank: bool = False) -> str:
    value = input(f"{prompt} [{default}]: ").strip()
    if value == "" and allow_blank:
        return ""
    return value or default


def prompt_yes_no(prompt: str, default: bool = False) -> bool:
    default_label = "Y/n" if default else "y/N"
    value = input(f"{prompt} [{default_label}]: ").strip().lower()
    if not value:
        return default
    return value in ("y", "yes")


def resolve_sheet_name(wb, requested_name: str) -> Optional[str]:
    if requested_name in wb.sheetnames:
        return requested_name
    normalized = requested_name.strip().lower()
    matches = [name for name in wb.sheetnames if name.strip().lower() == normalized]
    return matches[0] if matches else None


def parse_sheet_list(raw_value: str) -> list:
    return [item.strip() for item in raw_value.split(",") if item.strip()]


def parse_selection_numbers(raw_value: str, max_value: int) -> list:
    if not raw_value.strip():
        return []
    tokens = raw_value.replace(",", " ").split()
    selections = []
    for token in tokens:
        if not token.isdigit():
            continue
        value = int(token)
        if 1 <= value <= max_value:
            selections.append(value)
    seen = set()
    unique = []
    for value in selections:
        if value not in seen:
            seen.add(value)
            unique.append(value)
    return unique


def rgb_to_color(rgb: Optional[Any]) -> Optional[Dict[str, float]]:
    if not rgb:
        return None
    if hasattr(rgb, "rgb"):
        rgb = rgb.rgb
    elif hasattr(rgb, "value"):
        rgb = rgb.value
    if not isinstance(rgb, str):
        return None
    rgb = rgb.replace("0x", "").replace("#", "")
    if len(rgb) == 8:
        rgb = rgb[2:]  # Drop alpha channel
    if len(rgb) != 6:
        return None
    r = int(rgb[0:2], 16) / 255.0
    g = int(rgb[2:4], 16) / 255.0
    b = int(rgb[4:6], 16) / 255.0
    return {"red": r, "green": g, "blue": b}


def emu_to_pixels(emu: Optional[int]) -> Optional[int]:
    if not emu:
        return None
    return int(round(emu * 96 / 914400))


def build_number_format(cell) -> Optional[Dict[str, str]]:
    fmt = cell.number_format
    if not fmt:
        return None
    value = cell.value
    if isinstance(value, (datetime, date)) or is_date_format(fmt):
        nf_type = "DATE"
    elif "%" in fmt:
        nf_type = "PERCENT"
    elif fmt == "@":
        nf_type = "TEXT"
    else:
        nf_type = "NUMBER"
    return {"type": nf_type, "pattern": fmt}


def build_borders(cell) -> Optional[Dict[str, Dict[str, Any]]]:
    border = cell.border
    if not border:
        return None

    def convert_side(side):
        if not side or not side.style:
            return None
        style = BORDER_STYLE_MAP.get(side.style, "SOLID")
        color = rgb_to_color(getattr(side.color, "rgb", None))
        data = {"style": style}
        if color:
            data["color"] = color
        return data

    borders = {}
    for key, side in (("top", border.top), ("bottom", border.bottom),
                      ("left", border.left), ("right", border.right)):
        side_data = convert_side(side)
        if side_data:
            borders[key] = side_data
    return borders or None


def build_cell_format(cell) -> Optional[Dict[str, Any]]:
    fmt: Dict[str, Any] = {}

    if cell.font:
        text_format: Dict[str, Any] = {}
        if cell.font.bold:
            text_format["bold"] = True
        if cell.font.italic:
            text_format["italic"] = True
        if cell.font.underline:
            text_format["underline"] = True
        if cell.font.sz:
            text_format["fontSize"] = int(cell.font.sz)
        if cell.font.name:
            text_format["fontFamily"] = cell.font.name
        font_color = rgb_to_color(getattr(cell.font.color, "rgb", None))
        if font_color:
            text_format["foregroundColor"] = font_color
        if text_format:
            fmt["textFormat"] = text_format

    if cell.fill and cell.fill.patternType == "solid":
        fill_color = rgb_to_color(getattr(cell.fill.fgColor, "rgb", None))
        if fill_color:
            fmt["backgroundColor"] = fill_color

    if cell.alignment:
        if cell.alignment.horizontal:
            horiz = cell.alignment.horizontal.upper()
            horiz_map = {"CENTER_CONTINUOUS": "CENTER"}
            fmt["horizontalAlignment"] = horiz_map.get(horiz, horiz)
        if cell.alignment.vertical:
            vert = cell.alignment.vertical.upper()
            vert_map = {"CENTER": "MIDDLE", "JUSTIFY": "MIDDLE", "DISTRIBUTED": "MIDDLE"}
            fmt["verticalAlignment"] = vert_map.get(vert, vert)
        if cell.alignment.wrap_text:
            fmt["wrapStrategy"] = "WRAP"

    number_format = build_number_format(cell)
    if number_format:
        fmt["numberFormat"] = number_format

    borders = build_borders(cell)
    if borders:
        fmt["borders"] = borders

    return fmt or None


def build_cell_value(cell) -> Optional[Dict[str, Any]]:
    value = cell.value
    if value is None:
        return None
    if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
        return {"formulaValue": value}
    if isinstance(value, bool):
        return {"boolValue": value}
    if isinstance(value, (int, float)):
        return {"numberValue": float(value)}
    if isinstance(value, (datetime, date)):
        return {"numberValue": float(to_excel(value))}
    return {"stringValue": str(value)}


def build_rows(ws) -> Tuple[list, int, int]:
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    rows_data = []
    for r in range(1, max_row + 1):
        row_values = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell_data: Dict[str, Any] = {}
            value = build_cell_value(cell)
            if value:
                cell_data["userEnteredValue"] = value
            cell_format = build_cell_format(cell)
            if cell_format:
                cell_data["userEnteredFormat"] = cell_format
            row_values.append(cell_data)
        rows_data.append({"values": row_values})
    return rows_data, max_row, max_col


def convert_dimension_size_to_pixels(value: float, is_row: bool) -> Optional[int]:
    if value is None:
        return None
    if is_row:
        # Row height is in points
        return int(round(value * 96 / 72))
    # Column width in Excel units (~7 pixels per unit + 5 padding)
    return int(round(value * 7 + 5))


def build_dimension_requests(ws, sheet_id: int) -> list:
    requests = []

    for col_letter, dim in ws.column_dimensions.items():
        if dim.width:
            pixel_size = convert_dimension_size_to_pixels(dim.width, is_row=False)
            if pixel_size:
                start_col = column_index_from_string(col_letter) - 1
                requests.append({
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": sheet_id,
                            "dimension": "COLUMNS",
                            "startIndex": start_col,
                            "endIndex": start_col + 1,
                        },
                        "properties": {"pixelSize": pixel_size},
                        "fields": "pixelSize",
                    }
                })

    for row_idx, dim in ws.row_dimensions.items():
        if dim.height:
            pixel_size = convert_dimension_size_to_pixels(dim.height, is_row=True)
            if pixel_size:
                start_row = row_idx - 1
                requests.append({
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": sheet_id,
                            "dimension": "ROWS",
                            "startIndex": start_row,
                            "endIndex": start_row + 1,
                        },
                        "properties": {"pixelSize": pixel_size},
                        "fields": "pixelSize",
                    }
                })
    return requests


def build_merge_requests(ws, sheet_id: int) -> list:
    requests = []
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged))
        requests.append({
            "mergeCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": min_row - 1,
                    "endRowIndex": max_row,
                    "startColumnIndex": min_col - 1,
                    "endColumnIndex": max_col,
                },
                "mergeType": "MERGE_ALL",
            }
        })
    return requests


def build_conditional_format_requests(ws, sheet_id: int) -> list:
    requests = []
    cf_rules = ws.conditional_formatting._cf_rules
    dxfs = ws.parent._differential_styles.dxf

    for sqref, rules in cf_rules.items():
        min_col, min_row, max_col, max_row = range_boundaries(str(sqref.sqref))
        grid_range = {
            "sheetId": sheet_id,
            "startRowIndex": min_row - 1,
            "endRowIndex": max_row,
            "startColumnIndex": min_col - 1,
            "endColumnIndex": max_col,
        }
        for rule in rules:
            if rule.type != "expression" or not rule.formula:
                continue
            formula = rule.formula[0]
            if not formula.startswith("="):
                formula = "=" + formula
            dxf = dxfs[rule.dxfId] if rule.dxfId is not None else None
            bg_color = None
            if dxf and dxf.fill and dxf.fill.patternType == "solid":
                bg_color = rgb_to_color(getattr(dxf.fill.fgColor, "rgb", None))
            format_body = {}
            if bg_color:
                format_body["backgroundColor"] = bg_color
            requests.append({
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [grid_range],
                        "booleanRule": {
                            "condition": {
                                "type": "CUSTOM_FORMULA",
                                "values": [{"userEnteredValue": formula}],
                            },
                            "format": format_body,
                        },
                    },
                    "index": 0,
                }
            })
    return requests


def get_sheet_id_by_title(service, spreadsheet_id: str, title: str) -> Optional[int]:
    meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties").execute()
    for sheet in meta.get("sheets", []):
        props = sheet.get("properties", {})
        if props.get("title") == title:
            return props.get("sheetId")
    return None


def parse_range_reference(ref: str) -> Optional[Tuple[str, int, int, int, int]]:
    if not ref:
        return None
    if "!" not in ref:
        return None
    sheet_name, cell_range = ref.split("!", 1)
    sheet_name = sheet_name.strip("'")
    if ":" not in cell_range:
        cell_range = f"{cell_range}:{cell_range}"
    cell_range = cell_range.replace("$", "")
    start_cell, end_cell = cell_range.split(":", 1)
    start_col = column_index_from_string("".join([c for c in start_cell if c.isalpha()]))
    start_row = int("".join([c for c in start_cell if c.isdigit()]))
    end_col = column_index_from_string("".join([c for c in end_cell if c.isalpha()]))
    end_row = int("".join([c for c in end_cell if c.isdigit()]))
    return sheet_name, start_row, end_row, start_col, end_col


def get_chart_title(chart) -> str:
    try:
        return chart.title.tx.rich.p[0].r[0].t
    except Exception:
        return ""


def get_series_title_cell(series) -> Optional[Tuple[str, int, int, int, int]]:
    try:
        if series.title and series.title.strRef and series.title.strRef.f:
            return parse_range_reference(series.title.strRef.f)
    except Exception:
        return None
    return None


def get_series_value_range(series) -> Optional[Tuple[str, int, int, int, int]]:
    try:
        if series.val and series.val.numRef and series.val.numRef.f:
            return parse_range_reference(series.val.numRef.f)
    except Exception:
        return None
    return None


def get_series_category_range(series) -> Optional[Tuple[str, int, int, int, int]]:
    try:
        if series.cat:
            if series.cat.numRef and series.cat.numRef.f:
                return parse_range_reference(series.cat.numRef.f)
            if series.cat.strRef and series.cat.strRef.f:
                return parse_range_reference(series.cat.strRef.f)
    except Exception:
        return None
    return None


def build_grid_range(sheet_id: int, start_row: int, end_row: int, start_col: int, end_col: int) -> Dict[str, int]:
    return {
        "sheetId": sheet_id,
        "startRowIndex": start_row - 1,
        "endRowIndex": end_row,
        "startColumnIndex": start_col - 1,
        "endColumnIndex": end_col,
    }


def build_chart_position(chart, dest_sheet_id: int) -> Dict[str, Any]:
    anchor = getattr(chart, "anchor", None)
    row_index = 0
    col_index = 0
    width_pixels = None
    height_pixels = None
    try:
        row_index = anchor._from.row
        col_index = anchor._from.col
        width_pixels = emu_to_pixels(anchor.ext.cx) if anchor.ext else None
        height_pixels = emu_to_pixels(anchor.ext.cy) if anchor.ext else None
    except Exception:
        pass
    position = {
        "overlayPosition": {
            "anchorCell": {"sheetId": dest_sheet_id, "rowIndex": row_index, "columnIndex": col_index},
            "offsetXPixels": 0,
            "offsetYPixels": 0,
        }
    }
    if width_pixels:
        position["overlayPosition"]["widthPixels"] = width_pixels
    if height_pixels:
        position["overlayPosition"]["heightPixels"] = height_pixels
    return position


def build_line_chart_request(
    title: str,
    dest_sheet_id: int,
    position: Dict[str, Any],
    domain_range: Dict[str, int],
    series_ranges: list,
    x_axis_title: str,
    y_axis_title: str,
    header_count: int,
) -> Dict[str, Any]:
    return {
        "addChart": {
            "chart": {
                "spec": {
                    "title": title,
                    "basicChart": {
                        "chartType": "LINE",
                        "legendPosition": "RIGHT_LEGEND",
                        "axis": [
                            {"position": "LEFT_AXIS", "title": y_axis_title},
                            {"position": "BOTTOM_AXIS", "title": x_axis_title},
                        ],
                        "domains": [
                            {"domain": {"sourceRange": {"sources": [domain_range]}}}
                        ],
                        "series": [
                            {"series": {"sourceRange": {"sources": [sr]}}, "targetAxis": "LEFT_AXIS"}
                            for sr in series_ranges
                        ],
                        "headerCount": header_count,
                    },
                },
                "position": position,
            }
        }
    }


def get_sheet_properties_by_title(service, spreadsheet_id: str, title: str) -> Optional[Dict[str, Any]]:
    meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties").execute()
    for sheet in meta.get("sheets", []):
        props = sheet.get("properties", {})
        if props.get("title") == title:
            return props
    return None


def duplicate_sheet(service, spreadsheet_id: str, source_sheet_id: int, new_title: str) -> int:
    response = service.spreadsheets().sheets().copyTo(
        spreadsheetId=spreadsheet_id,
        sheetId=source_sheet_id,
        body={"destinationSpreadsheetId": spreadsheet_id},
    ).execute()
    new_sheet_id = response["sheetId"]
    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "requests": [
                {
                    "updateSheetProperties": {
                        "properties": {"sheetId": new_sheet_id, "title": new_title},
                        "fields": "title",
                    }
                }
            ]
        },
    ).execute()
    return new_sheet_id


def find_last_n_data_rows(values: list, n: int, start_row: int = 2, max_cols: int = 2) -> Optional[Tuple[int, int]]:
    rows_with_data = []
    for idx, row in enumerate(values, start=1):
        if idx < start_row:
            continue
        row_slice = row[:max_cols]
        if any(cell not in ("", None) for cell in row_slice):
            rows_with_data.append(idx)
    if not rows_with_data:
        return None
    selected = rows_with_data[-n:]
    return selected[0], selected[-1]


def add_last_8_weeks_net_sales_chart(
    service,
    spreadsheet_id: str,
    sheet_id: int,
    sheet_title: str,
) -> None:
    values = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_title}'!A:B",
    ).execute().get("values", [])
    row_range = find_last_n_data_rows(values, n=8, start_row=2, max_cols=2)
    if not row_range:
        print(f"Skipped chart for '{sheet_title}': no data found.")
        return
    start_row, end_row = row_range

    domain_range = build_grid_range(sheet_id, start_row, end_row, 1, 1)
    series_range = build_grid_range(sheet_id, start_row, end_row, 2, 2)
    position = {
        "overlayPosition": {
            "anchorCell": {"sheetId": sheet_id, "rowIndex": end_row + 1, "columnIndex": 0},
            "offsetXPixels": 0,
            "offsetYPixels": 0,
            "widthPixels": 850,
            "heightPixels": 360,
        }
    }
    chart_request = build_line_chart_request(
        "Net Sales (Last 8 Weeks)",
        sheet_id,
        position,
        domain_range,
        [series_range],
        "Week Ending",
        "Sales",
        header_count=0,
    )
    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id, body={"requests": [chart_request]}
    ).execute()


def copy_sheet_from_excel(
    service,
    spreadsheet_id: str,
    wb,
    source_sheet: str,
    dest_sheet: str,
    chart_template_sheet: str,
    build_charts: bool = True,
    allowed_chart_titles: Optional[set] = None,
) -> Optional[int]:
    resolved_sheet = resolve_sheet_name(wb, source_sheet)
    if not resolved_sheet:
        print(f"Sheet '{source_sheet}' not found in Excel workbook.")
        print(f"Available sheets: {', '.join(wb.sheetnames)}")
        retry_name = input("Enter the correct sheet name (blank to skip): ").strip()
        if not retry_name:
            print(f"Skipped '{source_sheet}'.")
            return
        resolved_sheet = resolve_sheet_name(wb, retry_name)
        if not resolved_sheet:
            print(f"Sheet '{retry_name}' not found. Skipping.")
            return
    source_sheet = resolved_sheet
    ws = wb[source_sheet]

    while True:
        existing_sheet_id = get_sheet_id_by_title(service, spreadsheet_id, dest_sheet)
        if existing_sheet_id is None:
            break
        print(f"Sheet '{dest_sheet}' already exists.")
        action = input("Choose: [o]verwrite, [r]ename, [s]kip: ").strip().lower()
        if action in ("o", "overwrite"):
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": [{"deleteSheet": {"sheetId": existing_sheet_id}}]},
            ).execute()
            break
        if action in ("r", "rename"):
            dest_sheet = input("Enter a new destination tab name: ").strip()
            if dest_sheet:
                continue
            print("Destination name cannot be blank.")
            continue
        if action in ("s", "skip"):
            print(f"Skipped '{source_sheet}'.")
            return None
        print("Invalid choice. Please enter o, r, or s.")

    new_sheet_id = None
    template_sheet_id = None
    template_sheet_props = None
    if chart_template_sheet:
        template_sheet_props = get_sheet_properties_by_title(service, spreadsheet_id, chart_template_sheet)
        if template_sheet_props:
            template_sheet_id = template_sheet_props.get("sheetId")

    if template_sheet_id is not None:
        new_sheet_id = duplicate_sheet(service, spreadsheet_id, template_sheet_id, dest_sheet)
    else:
        setup_response = service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": dest_sheet}}}]},
        ).execute()
        for reply in setup_response.get("replies", []):
            if "addSheet" in reply:
                new_sheet_id = reply["addSheet"]["properties"]["sheetId"]
        if new_sheet_id is None:
            raise SystemExit("Failed to create destination sheet.")

    rows_data, max_row, max_col = build_rows(ws)

    frozen_row_count = 0
    frozen_col_count = 0
    if template_sheet_props:
        grid_props = template_sheet_props.get("gridProperties", {})
        frozen_row_count = grid_props.get("frozenRowCount", 0) or 0
        frozen_col_count = grid_props.get("frozenColumnCount", 0) or 0
        if frozen_row_count or frozen_col_count:
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={
                    "requests": [
                        {
                            "updateSheetProperties": {
                                "properties": {
                                    "sheetId": new_sheet_id,
                                    "gridProperties": {
                                        "frozenRowCount": 0,
                                        "frozenColumnCount": 0,
                                    },
                                },
                                "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount",
                            }
                        }
                    ]
                },
            ).execute()

    requests = []
    requests.append({
        "updateSheetProperties": {
            "properties": {
                "sheetId": new_sheet_id,
                "gridProperties": {"rowCount": max_row, "columnCount": max_col},
            },
            "fields": "gridProperties(rowCount,columnCount)",
        }
    })
    requests.append({
        "updateCells": {
            "range": {
                "sheetId": new_sheet_id,
                "startRowIndex": 0,
                "startColumnIndex": 0,
                "endRowIndex": max_row,
                "endColumnIndex": max_col,
            },
            "rows": rows_data,
            "fields": "userEnteredValue,userEnteredFormat",
        }
    })

    requests.extend(build_dimension_requests(ws, new_sheet_id))
    requests.extend(build_merge_requests(ws, new_sheet_id))
    requests.extend(build_conditional_format_requests(ws, new_sheet_id))

    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id, body={"requests": requests}
    ).execute()

    if build_charts and not template_sheet_id:
        chart_requests = []
        title_axis_map = {
            "Net Sales (Last 8 Weeks)": ("Week Ending", "Sales"),
            "Sales vs Labor (Last 8 Weeks)": ("Week Ending", "Dollars"),
        }
        for chart in ws._charts:
            chart_title = get_chart_title(chart)
            if allowed_chart_titles and chart_title not in allowed_chart_titles:
                continue
            if chart_title not in title_axis_map:
                continue
            x_axis_title, y_axis_title = title_axis_map[chart_title]
            series_ranges = []
            domain_range = None
            header_count = 0
            for series in chart.series:
                val_ref = get_series_value_range(series)
                cat_ref = get_series_category_range(series)
                title_ref = get_series_title_cell(series)
                if not val_ref or not cat_ref:
                    continue
                sheet_name, start_row, end_row, start_col, end_col = val_ref
                cat_sheet, cat_start_row, cat_end_row, cat_start_col, cat_end_col = cat_ref
                if title_ref:
                    _, title_row, _, _, _ = title_ref
                    if title_row == start_row - 1:
                        start_row = title_row
                        header_count = 1
                if domain_range is None:
                    if title_ref and title_ref[1] == cat_start_row - 1:
                        cat_start_row = title_ref[1]
                        header_count = 1
                    data_sheet_id = get_sheet_id_by_title(service, spreadsheet_id, cat_sheet)
                    if data_sheet_id is None:
                        continue
                    domain_range = build_grid_range(data_sheet_id, cat_start_row, cat_end_row, cat_start_col, cat_end_col)
                data_sheet_id = get_sheet_id_by_title(service, spreadsheet_id, sheet_name)
                if data_sheet_id is None:
                    continue
                series_ranges.append(build_grid_range(data_sheet_id, start_row, end_row, start_col, end_col))

            if domain_range and series_ranges:
                position = build_chart_position(chart, new_sheet_id)
                chart_requests.append(
                    build_line_chart_request(
                        chart_title,
                        new_sheet_id,
                        position,
                        domain_range,
                        series_ranges,
                        x_axis_title,
                        y_axis_title,
                        header_count,
                    )
                )

        if chart_requests:
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id, body={"requests": chart_requests}
            ).execute()

    if frozen_row_count or frozen_col_count:
        try:
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={
                    "requests": [
                        {
                            "updateSheetProperties": {
                                "properties": {
                                    "sheetId": new_sheet_id,
                                    "gridProperties": {
                                        "frozenRowCount": frozen_row_count,
                                        "frozenColumnCount": frozen_col_count,
                                    },
                                },
                                "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount",
                            }
                        }
                    ]
                },
            ).execute()
        except HttpError as err:
            print(f"Warning: could not restore frozen panes: {err}")

    print(f"Copied '{source_sheet}' to Google Sheets tab '{dest_sheet}'.")
    return new_sheet_id


def main() -> None:
    config_path = "config.json"
    automation = CSVToSheetsAutomation(config_path=config_path, dry_run=True)
    if not automation.authenticate_google_sheets():
        raise SystemExit("Failed to authenticate to Google Sheets.")

    excel_path = automation.config.get("excel_file", "Restaurant_Daily_Ops_GSheets_Template_Targets.xlsx")
    if excel_path.endswith("_v4.xlsx"):
        excel_path = "Restaurant_Daily_Ops_GSheets_Template_Targets.xlsx"
    excel_path = prompt_with_default("Local Excel file path", excel_path)
    if not os.path.exists(excel_path):
        raise SystemExit(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path, data_only=False)
    spreadsheet_id = automation.config["google_sheet"]["sheet_id"]
    service = build("sheets", "v4", credentials=automation.creds)

    excluded_sheets = {
        "Sales_Payments",
        "Sales_Revenue",
        "Sales_Category",
        "Sales_Daypart",
        "Labor_Input",
        "COGS_Input",
        "Direct_NonControllable",
        "Targets_Reference",
        "COGS_Vendor_Lookup",
        "Job_Classification_Lookup",
    }
    selectable_sheets = [name for name in wb.sheetnames if name not in excluded_sheets]
    if not selectable_sheets:
        raise SystemExit("No selectable sheets found after exclusions.")

    print("Available Excel sheets:")
    for idx, sheet_name in enumerate(selectable_sheets, start=1):
        print(f"{idx}. {sheet_name}")
    raw_selection = input("Select sheets to copy by number (e.g., 1 3 5) [1]: ").strip()
    selections = parse_selection_numbers(raw_selection, len(selectable_sheets)) if raw_selection else [1]
    if not selections:
        raise SystemExit("No valid sheet selections provided.")

    for selection in selections:
        requested = selectable_sheets[selection - 1]
        resolved = resolve_sheet_name(wb, requested)
        source_sheet = resolved or requested
        dest_sheet = prompt_with_default(
            f"Destination Google Sheet tab name for '{source_sheet}'",
            source_sheet,
        )

        build_charts = False
        chart_template_sheet = ""
        allowed_chart_titles = None

        if source_sheet.strip().lower() == "dashboard":
            build_charts = True
            chart_template_sheet = prompt_with_default(
                "Google Sheet tab to clone charts from (blank to build from Excel)",
                source_sheet,
                allow_blank=True,
            )
        elif source_sheet.strip().lower() == "data_aggregation":
            build_charts = True
            allowed_chart_titles = {"Net Sales (Last 8 Weeks)"}

        sheet_id = copy_sheet_from_excel(
            service,
            spreadsheet_id,
            wb,
            source_sheet,
            dest_sheet,
            chart_template_sheet,
            build_charts=build_charts,
            allowed_chart_titles=allowed_chart_titles,
        )

        if sheet_id and source_sheet.strip().lower() == "data_aggregation":
            add_last_8_weeks_net_sales_chart(
                service,
                spreadsheet_id,
                sheet_id,
                dest_sheet,
            )


if __name__ == "__main__":
    main()
