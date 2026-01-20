"""
CSV to Google Sheets Automation Script
Reads daily CSV files from a subfolder and uploads them to Google Sheets 'daily ops' sheet.
"""

import pandas as pd
import os
import time
import json
import re
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple
import gspread
from google.oauth2.service_account import Credentials as ServiceAccountCredentials
from google.oauth2.credentials import Credentials as OAuthCredentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.auth.exceptions import GoogleAuthError
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
import pickle
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Fix Windows encoding issues - use ASCII-safe characters
CHECKMARK = "[OK]" if sys.platform == 'win32' else "{CHECKMARK}"
WARNING = "[!]" if sys.platform == 'win32' else "⚠"
CROSS = "[X]" if sys.platform == 'win32' else "{CROSS}"

# Default Google Drive folder IDs for production input sources
DRIVE_INPUT_ROOT_ID = "1fr63Bo6L7RBa3ID_shGWFM88SZI6xuPM"
DRIVE_SALES_INPUT_ID = "1JLIW_mG0xR-Ny0onSNt_QLTAuPsT7wDJ"
DRIVE_LABOR_INPUT_ID = "1JNHcL1SWNtp7ypQXsUWtCGp53z97UHzD"

class CSVToSheetsAutomation:
    def __init__(self, config_path: str = "config.json", dry_run: bool = False, process_oldest: bool = False, mode_override: Optional[bool] = None):
        """Initialize the automation with configuration file.
        
        Args:
            config_path: Path to configuration file
            dry_run: If True, validate but don't make changes
            process_oldest: If True, process oldest missing week instead of latest
            mode_override: If provided, override test_mode from config (True=testing, False=production)
        """
        self.config = self.load_config(config_path)
        self.csv_structure = self.load_csv_structure()
        self.dry_run = dry_run
        self.process_oldest = process_oldest
        # Use mode_override if provided, otherwise use config
        if mode_override is not None:
            self.test_mode = mode_override
        else:
            self.test_mode = self.config.get('test_mode', False)
        self.excel_file = self.config.get('excel_file', 'Restaurant_Daily_Ops_GSheets_Template_Targets.xlsx')
        self.excel_sheet_name = self.config.get('excel_sheet_name', 'Daily Ops')
        self.test_sheet_name = self.config.get('test_sheet_name', 'Category Daily Ops')
        self.auto_create_columns = self.config.get('auto_create_columns', False)
        self.test_process_csv_files = self.config.get('test_process_csv_files', [])  # Optional filter for test mode
        self.date_column_name = self.config.get('google_sheet', {}).get('date_column', 'Date')
        drive_inputs = self.config.get("drive_inputs", {})
        self.drive_input_root_id = drive_inputs.get("root_id", DRIVE_INPUT_ROOT_ID)
        self.drive_sales_input_id = drive_inputs.get("sales_input_id", DRIVE_SALES_INPUT_ID)
        self.drive_labor_input_id = drive_inputs.get("labor_input_id", DRIVE_LABOR_INPUT_ID)
        self.gc = None
        self.sheet = None
        self.worksheet = None
        self.workbook = None
        self.creds = None
        self.drive_service = None
        self._drive_sales_cache_path = None
        self._drive_labor_cache_path = None
        self.worksheet_cache = {}
        self.sheet_headers_cache = {}
        self.excel_worksheet = None
        self.excel_file_path = None
        self.using_test_sheet = False
        
        # CSV file to tab mapping for Sales Input processing
        self.csv_to_tab_mapping = {
            "Payments summary.csv": "Sales_Payments",
            "Revenue summary.csv": "Sales_Revenue",
            "Sales category summary.csv": "Sales_Category",
            "Service Daypart summary.csv": "Sales_Daypart"
        }
        
    def load_config(self, config_path: str) -> Dict:
        """Load configuration from JSON file and merge with secrets.json if available."""
        if not os.path.exists(config_path):
            raise FileNotFoundError(f"Configuration file {config_path} not found!")
        
        with open(config_path, 'r') as f:
            config = json.load(f)
        
        # Load secrets.json if it exists (contains sensitive data like Sheet ID and email)
        secrets_path = Path(__file__).parent / "secrets.json"
        if secrets_path.exists():
            try:
                with open(secrets_path, 'r', encoding='utf-8-sig') as f:
                    secrets = json.load(f)
                
                # Merge Sheet ID from secrets if provided
                if 'google_sheet_id' in secrets and secrets['google_sheet_id'] != 'YOUR_GOOGLE_SHEET_ID_HERE':
                    if 'google_sheet' not in config:
                        config['google_sheet'] = {}
                    config['google_sheet']['sheet_id'] = secrets['google_sheet_id']
                
                # Store email in config for use in error messages
                if 'email' in secrets and secrets['email'] != 'YOUR_EMAIL_HERE':
                    config['_email'] = secrets['email']
                
                # Store OAuth credentials if provided
                if 'oauth_credentials' in secrets:
                    config['_oauth_credentials'] = secrets['oauth_credentials']
                
                # Store Drive input folder IDs if provided
                if 'drive_inputs' in secrets and isinstance(secrets['drive_inputs'], dict):
                    config['drive_inputs'] = secrets['drive_inputs']
            except Exception as e:
                print(f"{WARNING} Warning: Could not load secrets.json: {e}")
        
        # Validate required config fields
        required_fields = ['google_sheet', 'csv_folder']
        for field in required_fields:
            if field not in config:
                raise ValueError(f"Missing required configuration field: {field}")
        
        return config
    
    def load_csv_structure(self) -> Dict:
        """Load CSV structure and mappings from csv_structure.json."""
        csv_structure_path = Path(__file__).parent / "csv_structure.json"
        
        if not csv_structure_path.exists():
            return {}
        
        try:
            with open(csv_structure_path, 'r', encoding='utf-8') as f:
                structure = json.load(f)
            return structure.get('csv_files', {})
        except Exception as e:
            return {}
    
    def authenticate_google_sheets(self) -> bool:
        """Authenticate with Google Sheets API using service account or OAuth."""
        auth_method = self.config.get('auth_method', 'service_account').lower()
        
        scope = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive.readonly"  # Needed to access sheets
        ]
        
        try:
            if auth_method == 'oauth':
                creds = self._authenticate_oauth(scope)
            else:
                creds = self._authenticate_service_account(scope)
            
            if not creds:
                return False
            
            self.creds = creds
            self.gc = gspread.authorize(creds)
            
            # Show which account is being used
            try:
                # Try to get user info from the credentials
                if hasattr(creds, 'id_token'):
                    # For OAuth, we can't easily get email from token, but we can try to list sheets
                    print(f"{CHECKMARK} Authenticated successfully")
                else:
                    print(f"{CHECKMARK} Authenticated successfully")
            except:
                pass
            
            # Open the Google Sheet
            sheet_id = self.config['google_sheet']['sheet_id']
            
            try:
                # Try opening by key first
                self.sheet = self.gc.open_by_key(sheet_id)
                print(f"{CHECKMARK} Successfully opened Google Sheet: {self.sheet.title}")
            except Exception as e:
                error_msg = str(e)
                print(f"\nError opening Google Sheet with ID {sheet_id[:20]}...")
                print(f"Error details: {error_msg}")
                
                # Check if this is the "not supported" error
                if "not supported" in error_msg.lower() or "400" in error_msg:
                    print(f"\n{WARNING} This error usually means:")
                    print(f"  1. OAuth app is in 'Testing' mode and your account isn't added as a test user")
                    print(f"  2. The OAuth app needs verification (for external users)")
                    print(f"  3. Domain restrictions on the Google Workspace account")
                    print(f"\nTrying alternative methods...")
                    
                    # Method 1: Try opening by URL
                    try:
                        sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}"
                        self.sheet = self.gc.open_by_url(sheet_url)
                        print(f"{CHECKMARK} Successfully opened Google Sheet via URL: {self.sheet.title}")
                    except Exception as e2:
                        print(f"  Method 1 (URL) failed: {e2}")
                        
                        # Method 2: Try listing all sheets and finding by ID
                        try:
                            print(f"  Trying Method 2: Listing accessible sheets...")
                            all_sheets = self.gc.openall()
                            print(f"  Found {len(all_sheets)} accessible sheet(s)")
                            
                            # Show first few sheets for debugging
                            print(f"  Sample of accessible sheets:")
                            for i, sheet in enumerate(all_sheets[:5], 1):
                                print(f"    {i}. {sheet.title} (ID: {sheet.id[:20]}...)")
                            
                            # Try to find by ID
                            found = False
                            for sheet in all_sheets:
                                if sheet.id == sheet_id:
                                    self.sheet = sheet
                                    print(f"{CHECKMARK} Found sheet by listing: {self.sheet.title}")
                                    found = True
                                    break
                            
                            if not found:
                                # Also try partial match (in case of formatting issues)
                                for sheet in all_sheets:
                                    if sheet.id.replace('-', '') == sheet_id.replace('-', ''):
                                        self.sheet = sheet
                                        print(f"{CHECKMARK} Found sheet by partial ID match: {self.sheet.title}")
                                        found = True
                                        break
                                
                                if not found:
                                    raise Exception(f"Sheet with ID {sheet_id} not found in {len(all_sheets)} accessible sheets")
                        except Exception as e3:
                            print(f"  Method 2 (listing) failed: {e3}")
                            print(f"\n💡 SOLUTION: Add your account as a test user in Google Cloud Console:")
                            print(f"  1. Go to: https://console.cloud.google.com/apis/credentials/consent")
                            print(f"  2. Click 'Edit App'")
                            print(f"  3. Go to 'Test users' section")
                            email = self.config.get('_email', 'YOUR_EMAIL_HERE')
                            print(f"  4. Click 'Add Users' and add: {email}")
                            print(f"  5. Save and re-run the script")
                
                if not hasattr(self, 'sheet') or not self.sheet:
                    print(f"\nTroubleshooting:")
                    print(f"  1. Make sure you authorized with the CORRECT Google account")
                    print(f"     - Use the account that OWNS or has EDIT access to the sheet")
                    email = self.config.get('_email', 'YOUR_EMAIL_HERE')
                    print(f"     - Based on your sheet, use: {email}")
                    print(f"  2. Verify the Sheet ID is correct: {sheet_id}")
                    print(f"  3. Check that the sheet is shared with the authorized account")
                    print(f"  4. Try opening the sheet in your browser to verify access")
                    print(f"  5. Delete token.pickle and re-run to re-authenticate with correct account")
                    return False
            
            # No worksheet needed - CSV processing functions access tabs directly by name
            self.worksheet = None
            
            return True
            
        except GoogleAuthError as e:
            print(f"\nAuthentication error: {e}")
            return False
        except Exception as e:
            print(f"\nError connecting to Google Sheets: {e}")
            return False
    
    def _authenticate_service_account(self, scope: List[str]):
        """Authenticate using service account credentials."""
        credentials_path = self.config.get('credentials_file', 'credentials.json')
        
        if not os.path.exists(credentials_path):
            print(f"\nError: Credentials file '{credentials_path}' not found!")
            print("Please follow the setup instructions in setup_instructions.md")
            return None
        
        try:
            creds = ServiceAccountCredentials.from_service_account_file(credentials_path, scopes=scope)
            return creds
        except Exception as e:
            print(f"\nService account authentication error: {e}")
            return None
    
    def _authenticate_oauth(self, scope: List[str]):
        """Authenticate using OAuth 2.0 flow."""
        token_path = self.config.get('oauth_token_file', 'token.pickle')
        
        # Check if OAuth credentials are in secrets.json first
        oauth_credentials = self.config.get('_oauth_credentials')
        credentials_path = None
        
        if not oauth_credentials:
            # Fall back to file-based credentials
            credentials_path = self.config.get('oauth_credentials_file', 'oauth_credentials.json')
            if not os.path.exists(credentials_path):
                print(f"\nError: OAuth credentials not found!")
                print("Please either:")
                print("  1. Add 'oauth_credentials' to secrets.json, or")
                print(f"  2. Create OAuth credentials file '{credentials_path}'")
                print("See setup_instructions.md for details")
                return None
        
        creds = None
        
        # Load existing token if available
        if os.path.exists(token_path):
            try:
                with open(token_path, 'rb') as token:
                    creds = pickle.load(token)
            except Exception as e:
                print(f"Warning: Could not load existing token: {e}")
        
        # If there are no (valid) credentials available, let the user log in
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                # Refresh the token
                try:
                    creds.refresh(Request())
                except Exception as e:
                    print(f"Error refreshing token: {e}")
                    creds = None
            
            if not creds:
                # Run the OAuth flow
                try:
                    if oauth_credentials:
                        # Use credentials from secrets.json
                        flow = InstalledAppFlow.from_client_config(oauth_credentials, scope)
                    else:
                        # Use credentials from file
                        flow = InstalledAppFlow.from_client_secrets_file(credentials_path, scope)
                    creds = flow.run_local_server(port=0)
                except Exception as e:
                    print(f"\nOAuth authentication error: {e}")
                    return None
            
            # Save the credentials for the next run
            try:
                with open(token_path, 'wb') as token:
                    pickle.dump(creds, token)
                print("{CHECKMARK} OAuth token saved for future use")
            except Exception as e:
                print(f"Warning: Could not save token: {e}")
        
        return creds
    
    def _get_drive_service(self):
        """Build and cache a Google Drive service client."""
        if self.drive_service:
            return self.drive_service
        if not self.creds:
            raise RuntimeError("Google credentials not initialized. Run authenticate_google_sheets first.")
        self.drive_service = build("drive", "v3", credentials=self.creds, cache_discovery=False)
        return self.drive_service

    def _retry_drive(self, func, *args, **kwargs):
        """Retry Google Drive calls when rate-limited."""
        max_retries = self.config.get('drive_max_retries', 5)
        base_delay = self.config.get('drive_retry_base_seconds', 5)

        last_exception = None
        for attempt in range(max_retries):
            try:
                return func(*args, **kwargs)
            except HttpError as e:
                last_exception = e
                status_code = getattr(e, "status_code", None)
                if status_code is None and hasattr(e, "resp"):
                    status_code = getattr(e.resp, "status", None)
                error_text = str(e)
                is_rate_limit = status_code in (429, 403) and (
                    "rateLimitExceeded" in error_text
                    or "userRateLimitExceeded" in error_text
                    or "quota" in error_text.lower()
                    or "429" in error_text
                )
                if not is_rate_limit:
                    raise
                delay = base_delay * (2 ** attempt)
                print(f"  {WARNING} Drive rate limit hit. Retrying in {delay} seconds...")
                time.sleep(delay)

        if last_exception:
            raise last_exception
        raise RuntimeError("Drive rate limit retries exhausted")
    
    def _list_drive_children(self, folder_id: str) -> List[Dict]:
        """List files and folders inside a Google Drive folder."""
        service = self._get_drive_service()
        items = []
        page_token = None
        
        while True:
            response = self._retry_drive(
                service.files().list(
                    q=f"'{folder_id}' in parents and trashed=false",
                    fields="nextPageToken, files(id, name, mimeType)",
                    pageToken=page_token
                ).execute
            )
            items.extend(response.get("files", []))
            page_token = response.get("nextPageToken")
            if not page_token:
                break
        
        return items
    
    def _download_drive_file(self, file_id: str, dest_path: Path) -> None:
        """Download a file from Google Drive to a local path."""
        service = self._get_drive_service()
        max_retries = self.config.get('drive_max_retries', 5)
        base_delay = self.config.get('drive_retry_base_seconds', 5)

        last_exception = None
        for attempt in range(max_retries):
            try:
                request = service.files().get_media(fileId=file_id)
                dest_path.parent.mkdir(parents=True, exist_ok=True)

                with open(dest_path, "wb") as f:
                    downloader = MediaIoBaseDownload(f, request)
                    done = False
                    while not done:
                        _, done = downloader.next_chunk()
                return
            except HttpError as e:
                last_exception = e
                if dest_path.exists():
                    try:
                        dest_path.unlink()
                    except OSError:
                        pass
                status_code = getattr(e, "status_code", None)
                if status_code is None and hasattr(e, "resp"):
                    status_code = getattr(e.resp, "status", None)
                error_text = str(e)
                is_rate_limit = status_code in (429, 403) and (
                    "rateLimitExceeded" in error_text
                    or "userRateLimitExceeded" in error_text
                    or "quota" in error_text.lower()
                    or "429" in error_text
                )
                if not is_rate_limit:
                    raise
                delay = base_delay * (2 ** attempt)
                print(f"  {WARNING} Drive rate limit hit. Retrying in {delay} seconds...")
                time.sleep(delay)

        if last_exception:
            raise last_exception
        raise RuntimeError("Drive download retries exhausted")
    
    def _sync_drive_folder_to_local(self, folder_id: str, dest_path: Path, recursive: bool = True, csv_only: bool = True) -> None:
        """Recursively download CSV files from a Drive folder into a local directory."""
        dest_path.mkdir(parents=True, exist_ok=True)
        items = self._list_drive_children(folder_id)
        
        for item in items:
            name = item["name"]
            mime_type = item["mimeType"]
            item_id = item["id"]
            
            if mime_type == "application/vnd.google-apps.folder":
                if recursive:
                    self._sync_drive_folder_to_local(item_id, dest_path / name, recursive=recursive, csv_only=csv_only)
                continue
            
            if csv_only and not name.lower().endswith(".csv"):
                continue
            
            self._download_drive_file(item_id, dest_path / name)
    
    def _prepare_drive_cache_dir(self) -> Path:
        """Get local cache directory for Drive downloads."""
        return Path(__file__).parent / "_drive_cache"
    
    def _prepare_sales_input_folder_from_drive(self, folder_id: Optional[str] = None, folder_name: Optional[str] = None) -> Path:
        """Download Sales_Input folder (or a specific subfolder) from Google Drive into local cache."""
        if self._drive_sales_cache_path and self._drive_sales_cache_path.exists():
            if folder_id and folder_name:
                target_path = self._drive_sales_cache_path / folder_name
                if target_path.exists():
                    return target_path
            else:
                return self._drive_sales_cache_path
        
        cache_root = self._prepare_drive_cache_dir()
        sales_cache = cache_root / "Sales_Input"
        sales_cache.mkdir(parents=True, exist_ok=True)

        if folder_id and folder_name:
            target_path = sales_cache / folder_name
            if target_path.exists():
                shutil.rmtree(target_path)
            print(f"  Downloading {folder_name} from Google Drive...")
            self._sync_drive_folder_to_local(folder_id, target_path, recursive=True, csv_only=True)
            self._drive_sales_cache_path = sales_cache
            return target_path

        if sales_cache.exists():
            shutil.rmtree(sales_cache)
        sales_cache.mkdir(parents=True, exist_ok=True)
        
        print("  Downloading Sales_Input from Google Drive...")
        self._sync_drive_folder_to_local(self.drive_sales_input_id, sales_cache, recursive=True, csv_only=True)
        
        self._drive_sales_cache_path = sales_cache
        return sales_cache
    
    def _prepare_labor_input_folder_from_drive(self) -> Path:
        """Download Labor_Input folder from Google Drive into local cache."""
        if self._drive_labor_cache_path and self._drive_labor_cache_path.exists():
            return self._drive_labor_cache_path
        
        cache_root = self._prepare_drive_cache_dir()
        labor_cache = cache_root / "Labor_Input"
        if labor_cache.exists():
            shutil.rmtree(labor_cache)
        labor_cache.mkdir(parents=True, exist_ok=True)
        
        print("  Downloading Labor_Input from Google Drive...")
        self._sync_drive_folder_to_local(self.drive_labor_input_id, labor_cache, recursive=False, csv_only=True)
        
        self._drive_labor_cache_path = labor_cache
        return labor_cache

    def _retry_gspread(self, func, *args, **kwargs):
        """Retry Google Sheets calls when rate-limited (HTTP 429)."""
        max_retries = self.config.get('gspread_max_retries', 5)
        base_delay = self.config.get('gspread_retry_base_seconds', 5)

        last_exception = None
        for attempt in range(max_retries):
            try:
                return func(*args, **kwargs)
            except gspread.exceptions.APIError as e:
                last_exception = e
                response = getattr(e, "response", None)
                status_code = getattr(response, "status_code", None) or getattr(response, "status", None)
                is_rate_limit = status_code == 429 or "429" in str(e)
                if not is_rate_limit:
                    raise
                delay = base_delay * (2 ** attempt)
                print(f"  {WARNING} Rate limit hit. Retrying in {delay} seconds...")
                time.sleep(delay)

        if last_exception:
            raise last_exception
        raise RuntimeError("Rate limit retries exhausted")

    def _get_worksheet(self, tab_name: str):
        """Get worksheet by name with caching and retry."""
        if tab_name in self.worksheet_cache:
            return self.worksheet_cache[tab_name]
        if not self.sheet:
            raise RuntimeError("Google Sheet not loaded")
        worksheet = self._retry_gspread(self.sheet.worksheet, tab_name)
        self.worksheet_cache[tab_name] = worksheet
        return worksheet

    def _get_sheet_headers(self, tab_name: str, worksheet=None, force_refresh: bool = False) -> List[str]:
        """Get cached header row values for a worksheet."""
        if not force_refresh and tab_name in self.sheet_headers_cache:
            return self.sheet_headers_cache[tab_name]
        ws = worksheet or self._get_worksheet(tab_name)
        if not ws:
            return []
        headers = self._retry_gspread(ws.row_values, 1)
        self.sheet_headers_cache[tab_name] = headers
        return headers
    
    def get_csv_folder_path(self) -> Path:
        """Get the path to the CSV folder. Auto-detects dated folders (e.g., SalesSummary_2025-12-31_2025-12-31).
        Selects the folder with the latest date in its name, or oldest missing week if process_oldest is True."""
        # Prepare the base folder (production or test mode)
        if not self.test_mode:
            folder_path = self._prepare_sales_input_folder_from_drive()
        else:
            csv_folder = self.config['csv_folder']
            base_path = Path(__file__).parent
            
            # If absolute path, use it; otherwise relative to script location
            if os.path.isabs(csv_folder):
                folder_path = Path(csv_folder)
            else:
                folder_path = base_path / csv_folder
        
        # If the exact folder exists and contains CSV files directly, use it
        if folder_path.exists() and folder_path.is_dir():
            csv_files = list(folder_path.glob("*.csv"))
            if csv_files:
                return folder_path
        
        # If process_oldest is True, try to find oldest missing week (works for both test_mode and production)
        if self.process_oldest:
            oldest_missing = self.find_oldest_missing_sales_folder()
            if oldest_missing:
                folder, week_ending_date = oldest_missing
                week_ending_str = week_ending_date.strftime("%Y-%m-%d")
                print(f"  Found oldest missing week: {week_ending_str} in folder: {folder.name}")
                return folder
        
        # Otherwise, look for folders matching the pattern (e.g., SalesSummary_2025-12-31_2025-12-31 or daily_data_01_07_2025)
        # Look inside the configured folder (not parent) for subfolders starting with "SalesSummary"
        search_dir = folder_path if folder_path.exists() and folder_path.is_dir() else folder_path.parent
        
        # Find folders that start with "SalesSummary" (for new pattern)
        matching_folders = []
        if search_dir.exists() and search_dir.is_dir():
            matching_folders = [d for d in search_dir.iterdir() 
                               if d.is_dir() and d.name.startswith("SalesSummary")]
        
        if matching_folders:
            # Extract dates from folder names and select the one with the latest date
            folders_with_dates = []
            for folder in matching_folders:
                date = self._extract_date_from_string(folder.name)
                if date:
                    folders_with_dates.append((folder, date))
            
            if folders_with_dates:
                # Sort by date (latest first)
                folders_with_dates.sort(key=lambda x: x[1], reverse=True)
                selected_folder = folders_with_dates[0][0]
                return selected_folder
            else:
                # No dates found in names, use most recently modified
                matching_folders.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                selected_folder = matching_folders[0]
                print(f"  Found folder: {selected_folder.name} (no date in name, using most recent)")
                return selected_folder
        
        # Return the original path if no match found (will be created if needed)
        return folder_path
    
    def _extract_date_from_string(self, text: str) -> Optional[datetime]:
        """Helper method to extract date from a string (used for folder names)."""
        # Pattern: SalesSummary_YYYY-MM-DD_YYYY-MM-DD (e.g., SalesSummary_2025-12-31_2025-12-31)
        date_patterns = [
            r'SalesSummary_(\d{4})-(\d{2})-(\d{2})_\d{4}-\d{2}-\d{2}',  # SalesSummary_2025-12-31_2025-12-31
            # Legacy patterns for backward compatibility
            r'daily_data_(\d{2})_(\d{2})_(\d{4})',  # daily_data_01_07_2025
            r'daily_data_(\d{2})-(\d{2})-(\d{4})',  # daily_data_01-07-2025
            r'daily_data(\d{8})',  # daily_data01072025
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, text)
            if match:
                if pattern.startswith(r'SalesSummary'):
                    # SalesSummary_YYYY-MM-DD_YYYY-MM-DD format
                    year, month, day = match.groups()
                    try:
                        return datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d")
                    except ValueError:
                        continue
                elif len(match.groups()) == 3:
                    # MM_DD_YYYY or MM-DD-YYYY format (legacy)
                    month, day, year = match.groups()
                    try:
                        return datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d")
                    except ValueError:
                        continue
                elif len(match.groups()) == 1:
                    # YYYYMMDD format (legacy)
                    date_str = match.group(1)
                    try:
                        return datetime.strptime(date_str, "%Y%m%d")
                    except ValueError:
                        continue
        
        return None
    
    def extract_date_from_folder_name(self) -> Optional[datetime]:
        """Extract date from folder name (e.g., SalesSummary_2025-12-31_2025-12-31 -> 2025-12-31 or daily_data_01_07_2025 -> 2025-01-07)."""
        csv_folder = self.get_csv_folder_path()
        folder_name = csv_folder.name
        return self._extract_date_from_string(folder_name)
    
    def extract_week_ending_date(self) -> Optional[datetime]:
        """Extract input date (second date) from folder name (e.g., SalesSummary_2025-12-29_2026-01-04 -> 2026-01-04)."""
        csv_folder = self.get_csv_folder_path()
        folder_name = csv_folder.name
        
        # Pattern: SalesSummary_YYYY-MM-DD_YYYY-MM-DD (extract second date - input date)
        pattern = r'SalesSummary_\d{4}-\d{2}-\d{2}_(\d{4})-(\d{2})-(\d{2})'
        match = re.search(pattern, folder_name)
        
        if match:
            year, month, day = match.groups()
            try:
                return datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d")
            except ValueError:
                pass
        
        return None
    
    def extract_week_ending_date_from_payroll_export(self, csv_file: Path) -> Optional[datetime]:
        """Extract input date from PayrollExport CSV filename.
        
        Supports two formats:
        - PayrollExport_YYYY_MM_DD.csv (single date)
        - PayrollExport_YYYY_MM_DD-YYYY_MM_DD.csv (extract second date - input date)
        Examples: 
          PayrollExport_2025_12_30.csv -> 2025-12-30
          PayrollExport_2025_12_29-2026_01_04.csv -> 2026-01-04
        """
        filename = csv_file.name
        
        # First try: PayrollExport_YYYY_MM_DD-YYYY_MM_DD (extract second date - input date)
        pattern_two_dates = r'PayrollExport_\d{4}_\d{2}_\d{2}-(\d{4})_(\d{2})_(\d{2})'
        match = re.search(pattern_two_dates, filename)
        
        if match:
            year, month, day = match.groups()
            try:
                return datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d")
            except ValueError:
                pass
        
        # Second try: PayrollExport_YYYY_MM_DD (single date format)
        pattern_single_date = r'PayrollExport_(\d{4})_(\d{2})_(\d{2})\.csv'
        match = re.search(pattern_single_date, filename)
        
        if match:
            year, month, day = match.groups()
            try:
                return datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d")
            except ValueError:
                pass
        
        return None
    
    def find_latest_labor_input_csv(self, labor_input_folder: Path) -> Tuple[Optional[Path], Optional[datetime], List[Path]]:
        """Find the latest PayrollExport CSV file in Labor_Input folder.
        
        Returns:
            Tuple of (latest_csv_file, week_ending_date, duplicate_files)
            - latest_csv_file: The most recent CSV file by modification time
            - week_ending_date: Week ending date extracted from filename
            - duplicate_files: List of other CSV files with the same input date (if any)
        """
        if not labor_input_folder.exists():
            print(f"  {CROSS} Labor_Input folder does not exist: {labor_input_folder}")
            return None, None, []
        
        # Find all PayrollExport CSV files
        csv_files = list(labor_input_folder.glob("PayrollExport_*.csv"))
        
        if not csv_files:
            print(f"  {WARNING} No PayrollExport CSV files found in {labor_input_folder}")
            return None, None, []
        
        # Sort by modification time (newest first)
        csv_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        
        # Extract input dates and group by date
        file_date_map = {}
        for csv_file in csv_files:
            week_ending_date = self.extract_week_ending_date_from_payroll_export(csv_file)
            if week_ending_date:
                week_ending_str = week_ending_date.strftime("%Y-%m-%d")
                if week_ending_str not in file_date_map:
                    file_date_map[week_ending_str] = []
                file_date_map[week_ending_str].append(csv_file)
        
        if not file_date_map:
            print(f"  {WARNING} Could not extract input dates from any CSV files")
            return None, None, []
        
        # Get the latest file (first in sorted list)
        latest_file = csv_files[0]
        latest_week_ending = self.extract_week_ending_date_from_payroll_export(latest_file)
        
        if not latest_week_ending:
            print(f"  {WARNING} Could not extract input date from latest file: {latest_file.name}")
            return None, None, []
        
        latest_week_ending_str = latest_week_ending.strftime("%Y-%m-%d")
        
        # Check for duplicate files with same input date
        duplicate_files = file_date_map[latest_week_ending_str]
        if len(duplicate_files) > 1:
            # Exclude the latest file from duplicates list
            duplicate_files = [f for f in duplicate_files if f != latest_file]
            return latest_file, latest_week_ending, duplicate_files
        
        return latest_file, latest_week_ending, []
    
    def get_all_existing_week_ending_dates(self, tab_name: str) -> "Set[str]":
        """Get all existing dates from a tab.
        Returns a set of date strings in YYYY-MM-DD format.
        Works with both Excel (test_mode) and Google Sheets (production)."""
        existing_dates = set()
        
        if self.test_mode:
            # Excel version
            if not self.workbook or tab_name not in self.workbook.sheetnames:
                return existing_dates
            
            worksheet = self.workbook[tab_name]
            
            if worksheet.max_row == 0:
                return existing_dates
            
            # Check all data rows (starting from row 2)
            for row_idx in range(2, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row_idx, column=1).value
                if cell_value is not None:
                    # Convert to string for comparison
                    try:
                        if isinstance(cell_value, datetime):
                            date_str = cell_value.strftime("%Y-%m-%d")
                        elif isinstance(cell_value, pd.Timestamp):
                            date_str = cell_value.strftime("%Y-%m-%d")
                        else:
                            date_str = pd.to_datetime(str(cell_value)).strftime("%Y-%m-%d")
                        existing_dates.add(date_str)
                    except:
                        pass
        else:
            # Google Sheets version
            if not self.sheet:
                return existing_dates
            
            try:
                worksheet = self._get_worksheet(tab_name)
            except gspread.exceptions.WorksheetNotFound:
                return existing_dates
            
            # Get all values from the first column (Date column)
            try:
                all_values = self._retry_gspread(worksheet.col_values, 1)  # Column A (index 1)
                if len(all_values) <= 1:  # Only header or empty
                    return existing_dates
                
                # Check all data rows (starting from row 2, index 1)
                for cell_value in all_values[1:]:  # Skip header
                    if cell_value:
                        # Convert to string for comparison
                        try:
                            cell_str = pd.to_datetime(str(cell_value)).strftime("%Y-%m-%d")
                            existing_dates.add(cell_str)
                        except:
                            pass
            except Exception as e:
                return existing_dates
        
        return existing_dates
    
    def find_all_sales_folders_with_dates(self) -> List[Tuple[Path, datetime]]:
        """Find all SalesSummary folders with their input dates.
        Returns a list of (folder_path, input_date) tuples, sorted by date (oldest first)."""
        if self.test_mode:
            csv_folder = self.config['csv_folder']
            base_path = Path(__file__).parent
            
            # If absolute path, use it; otherwise relative to script location
            if os.path.isabs(csv_folder):
                folder_path = Path(csv_folder)
            else:
                folder_path = base_path / csv_folder
        else:
            # Production mode pulls Sales_Input from Drive into local cache
            folder_path = self.get_csv_folder_path()
        
        # Look for folders matching the pattern
        search_dir = folder_path if folder_path.exists() and folder_path.is_dir() else folder_path.parent
        
        matching_folders = []
        if search_dir.exists() and search_dir.is_dir():
            matching_folders = [d for d in search_dir.iterdir() 
                               if d.is_dir() and d.name.startswith("SalesSummary")]
        
        folders_with_dates = []
        for folder in matching_folders:
            week_ending_date = self.extract_week_ending_date_from_folder(folder)
            if week_ending_date:
                folders_with_dates.append((folder, week_ending_date))
        
        # Sort by date (oldest first)
        folders_with_dates.sort(key=lambda x: x[1])
        return folders_with_dates

    def _extract_input_date_from_sales_folder_name(self, folder_name: str) -> Optional[datetime]:
        """Extract input date (second date) from a SalesSummary folder name."""
        pattern = r'SalesSummary_\d{4}-\d{2}-\d{2}_(\d{4})-(\d{2})-(\d{2})'
        match = re.search(pattern, folder_name)
        if not match:
            return None
        year, month, day = match.groups()
        try:
            return datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d")
        except ValueError:
            return None

    def find_all_sales_drive_folders_with_dates(self) -> List[Tuple[str, str, datetime]]:
        """Find all SalesSummary folders in Drive with their input dates."""
        items = self._list_drive_children(self.drive_sales_input_id)
        folders_with_dates = []
        for item in items:
            if item.get("mimeType") != "application/vnd.google-apps.folder":
                continue
            folder_name = item.get("name", "")
            if not folder_name.startswith("SalesSummary"):
                continue
            input_date = self._extract_input_date_from_sales_folder_name(folder_name)
            if input_date:
                folders_with_dates.append((item["id"], folder_name, input_date))
        
        folders_with_dates.sort(key=lambda x: x[2])
        return folders_with_dates

    def find_missing_sales_folders(self) -> List[Tuple[Path, datetime]]:
        """Find all SalesSummary folders whose input dates are missing in any sales tab."""
        existing_by_tab = {}
        for tab_name in self.csv_to_tab_mapping.values():
            existing_by_tab[tab_name] = self.get_all_existing_week_ending_dates(tab_name)

        folders_with_dates = self.find_all_sales_folders_with_dates()
        if not folders_with_dates:
            return []

        missing = []
        for folder, input_date in folders_with_dates:
            input_date_str = input_date.strftime("%Y-%m-%d")
            if any(input_date_str not in existing for existing in existing_by_tab.values()):
                missing.append((folder, input_date))

        return missing

    def find_missing_sales_drive_folders(self) -> List[Tuple[str, str, datetime]]:
        """Find all SalesSummary Drive folders whose input dates are missing in any sales tab."""
        existing_by_tab = {}
        for tab_name in self.csv_to_tab_mapping.values():
            existing_by_tab[tab_name] = self.get_all_existing_week_ending_dates(tab_name)

        folders_with_dates = self.find_all_sales_drive_folders_with_dates()
        if not folders_with_dates:
            return []

        missing = []
        for folder_id, folder_name, input_date in folders_with_dates:
            input_date_str = input_date.strftime("%Y-%m-%d")
            if any(input_date_str not in existing for existing in existing_by_tab.values()):
                missing.append((folder_id, folder_name, input_date))

        return missing
    
    def extract_week_ending_date_from_folder(self, folder: Path) -> Optional[datetime]:
        """Extract input date (second date) from folder name."""
        folder_name = folder.name
        # Pattern: SalesSummary_YYYY-MM-DD_YYYY-MM-DD (extract second date - input date)
        pattern = r'SalesSummary_\d{4}-\d{2}-\d{2}_(\d{4})-(\d{2})-(\d{2})'
        match = re.search(pattern, folder_name)
        
        if match:
            year, month, day = match.groups()
            try:
                return datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d")
            except ValueError:
                pass
        
        return None
    
    def find_oldest_missing_sales_folder(self) -> Optional[Tuple[Path, datetime]]:
        """Find the oldest SalesSummary folder whose input date is missing in any sales tab.
        Returns (folder_path, input_date) or None if all dates exist or no folders found.
        Works with both Excel (test_mode) and Google Sheets (production)."""
        missing = self.find_missing_sales_folders()
        if not missing:
            return None
        return missing[0]
    
    def find_oldest_missing_labor_csv(self, labor_input_folder: Path) -> Optional[Tuple[Path, datetime]]:
        """Find the oldest PayrollExport CSV file whose input date doesn't exist in Labor_Input tab.
        Returns (csv_file, input_date) or None if all dates exist or no files found.
        Works with both Excel (test_mode) and Google Sheets (production)."""
        if not labor_input_folder.exists():
            return None
        
        # Find all PayrollExport CSV files
        csv_files = list(labor_input_folder.glob("PayrollExport_*.csv"))
        
        if not csv_files:
            return None
        
        # Get all existing dates from Labor_Input tab
        existing_dates = self.get_all_existing_week_ending_dates("Labor_Input")
        
        # Extract input dates and sort by date (oldest first)
        files_with_dates = []
        for csv_file in csv_files:
            input_date = self.extract_week_ending_date_from_payroll_export(csv_file)
            if input_date:
                files_with_dates.append((csv_file, input_date))
        
        if not files_with_dates:
            return None
        
        # Sort by date (oldest first)
        files_with_dates.sort(key=lambda x: x[1])
        
        # Find the oldest file whose input date doesn't exist
        for csv_file, input_date in files_with_dates:
            input_date_str = input_date.strftime("%Y-%m-%d")
            if input_date_str not in existing_dates:
                return (csv_file, input_date)
        
        return None  # All weeks already exist

    def find_missing_labor_csvs(self, labor_input_folder: Path) -> List[Tuple[Path, datetime, List[Path]]]:
        """Find all PayrollExport CSV files whose input dates don't exist in Labor_Input tab.
        Returns list of (selected_file, input_date, duplicate_files) sorted by date (oldest first)."""
        if not labor_input_folder.exists():
            return []

        csv_files = list(labor_input_folder.glob("PayrollExport_*.csv"))
        if not csv_files:
            return []

        existing_dates = self.get_all_existing_week_ending_dates("Labor_Input")

        # Group by input date, track duplicates
        file_date_map = {}
        for csv_file in csv_files:
            input_date = self.extract_week_ending_date_from_payroll_export(csv_file)
            if input_date:
                input_date_str = input_date.strftime("%Y-%m-%d")
                file_date_map.setdefault(input_date_str, []).append(csv_file)

        missing = []
        for input_date_str, files in file_date_map.items():
            if input_date_str in existing_dates:
                continue
            # Choose latest file by modification time for this date
            files_sorted = sorted(files, key=lambda x: x.stat().st_mtime, reverse=True)
            selected_file = files_sorted[0]
            duplicate_files = files_sorted[1:]
            input_date = datetime.strptime(input_date_str, "%Y-%m-%d")
            missing.append((selected_file, input_date, duplicate_files))

        # Sort by date (oldest first)
        missing.sort(key=lambda x: x[1])
        return missing
    
    def find_csv_files(self, folder_path: Optional[Path] = None) -> List[Path]:
        """Find all CSV files in the provided folder (or configured folder if None).
        Filters by test_process_csv_files if configured."""
        csv_folder = folder_path or self.get_csv_folder_path()
        
        if not csv_folder.exists():
            print(f"CSV folder '{csv_folder}' does not exist. Creating it...")
            csv_folder.mkdir(parents=True, exist_ok=True)
            print(f"Please download your CSV files to: {csv_folder}")
            return []
        
        csv_files = list(csv_folder.glob("*.csv"))
        
        if not csv_files:
            print(f"No CSV files found in {csv_folder}")
            return []
        
        # Filter CSV files if test_process_csv_files is configured (in test mode)
        if self.test_mode and self.test_process_csv_files:
            filtered_files = []
            allowed_files = set(self.test_process_csv_files)
            
            for csv_file in csv_files:
                if csv_file.name in allowed_files:
                    filtered_files.append(csv_file)
                else:
                    print(f"  [SKIP] {csv_file.name} (not in test_process_csv_files list)")
            
            csv_files = filtered_files
            
            if not csv_files:
                print(f"No CSV files match the filter: {self.test_process_csv_files}")
                return []
        
        return csv_files
    
    def _read_csv_safe(self, csv_file: Path) -> Optional[pd.DataFrame]:
        """Read a CSV file safely, returning None if empty or unreadable."""
        if not csv_file.exists():
            print(f"  {WARNING} CSV file not found: {csv_file}")
            return None
        try:
            if csv_file.stat().st_size == 0:
                print(f"  {WARNING} CSV file is empty: {csv_file.name}")
                return None
        except OSError as e:
            print(f"  {WARNING} Could not read CSV file size: {csv_file.name} ({e})")
            return None

        try:
            return pd.read_csv(csv_file)
        except pd.errors.EmptyDataError:
            print(f"  {WARNING} CSV file has no data: {csv_file.name}")
            return None
        except Exception as e:
            print(f"  {WARNING} Failed to read CSV file {csv_file.name}: {e}")
            return None

    def extract_date_from_csv(self, csv_file: Path) -> Optional[datetime]:
        """Extract date from CSV file content."""
        try:
            df = self._read_csv_safe(csv_file)
            if df is None or df.empty:
                return None
            
            # Check for date column in various formats
            date_columns = [col for col in df.columns 
                          if 'date' in col.lower() or 'yyyyMMdd' in col or 'day' in col.lower()]
            
            if date_columns:
                date_col = date_columns[0]
                first_date = df[date_col].iloc[0]
                
                # Handle yyyyMMdd format (e.g., 20260106)
                if isinstance(first_date, (int, float)) and len(str(int(first_date))) == 8:
                    date_str = str(int(first_date))
                    try:
                        return datetime.strptime(date_str, "%Y%m%d")
                    except ValueError:
                        pass
                
                # Try parsing as date string
                if isinstance(first_date, str):
                    # Try common date formats
                    date_formats = ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y%m%d"]
                    for fmt in date_formats:
                        try:
                            return datetime.strptime(first_date, fmt)
                        except ValueError:
                            continue
            
            # Check filename for date pattern
            filename = csv_file.stem
            date_patterns = [
                r'(\d{4}-\d{2}-\d{2})',  # 2026-01-06
                r'(\d{8})',  # 20260106
                r'(\d{2}/\d{2}/\d{4})',  # 01/06/2026
            ]
            
            for pattern in date_patterns:
                match = re.search(pattern, filename)
                if match:
                    date_str = match.group(1)
                    # Try parsing
                    if len(date_str) == 8 and date_str.isdigit():
                        try:
                            return datetime.strptime(date_str, "%Y%m%d")
                        except ValueError:
                            pass
                    else:
                        date_formats = ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]
                        for fmt in date_formats:
                            try:
                                return datetime.strptime(date_str, fmt)
                            except ValueError:
                                continue
            
            print(f"  Warning: Could not extract date from {csv_file.name}")
            return None
            
        except Exception as e:
            print(f"  Error reading {csv_file.name}: {e}")
            return None
    
    def get_date_column_index(self) -> Optional[int]:
        """Get the index of the date column in the Google Sheet."""
        date_col_name = self.config['google_sheet'].get('date_column')
        if not date_col_name:
            return None
        
        # Get header row (assuming row 1)
        headers = self._retry_gspread(self.worksheet.row_values, 1)
        
        try:
            # Try exact match first
            col_index = headers.index(date_col_name) + 1  # +1 for 1-based indexing
            return col_index
        except ValueError:
            # Try case-insensitive match
            try:
                col_index = next(i for i, h in enumerate(headers, 1) if h.lower() == date_col_name.lower())
                print(f"  Note: Found date column '{headers[col_index-1]}' (case-insensitive match)")
                return col_index
            except StopIteration:
                print(f"Warning: Date column '{date_col_name}' not found in sheet headers")
                print(f"  Available headers: {headers[:15]}...")  # Show first 15 headers
                return None
    
    def find_row_for_date(self, target_date: datetime) -> Optional[int]:
        """Find the row number for a given date in target (Google Sheets or Excel)."""
        if self.test_mode:
            # For Excel, find existing row
            excel_path = Path(__file__).parent / self.excel_file
            try:
                df = pd.read_excel(excel_path, sheet_name=self.excel_sheet_name, header=0)
                date_column = self.config['google_sheet'].get('date_column', 'Date')
                
                if date_column not in df.columns:
                    return None
                
                target_date_str = target_date.strftime("%Y-%m-%d")
                
                for idx, row_date in enumerate(df[date_column], start=2):
                    if pd.notna(row_date):
                        if isinstance(row_date, pd.Timestamp):
                            row_date_str = row_date.strftime("%Y-%m-%d")
                        elif isinstance(row_date, datetime):
                            row_date_str = row_date.strftime("%Y-%m-%d")
                        elif isinstance(row_date, str):
                            try:
                                row_date_str = pd.to_datetime(row_date).strftime("%Y-%m-%d")
                            except:
                                row_date_str = str(row_date).strip()
                        else:
                            row_date_str = str(row_date).strip()
                        
                        if row_date_str == target_date_str:
                            return idx
                
                return None  # Row doesn't exist yet, will be created by create_row_for_date
            except Exception as e:
                return None
        
        # Google Sheets implementation (existing code)
        date_col_index = self.get_date_column_index()
        if not date_col_index:
            return None
        
        # Get all values in the date column
        date_col = self._retry_gspread(self.worksheet.col_values, date_col_index)
        
        # Format target date for comparison
        target_date_strs = [
            target_date.strftime("%Y-%m-%d"),
            target_date.strftime("%m/%d/%Y"),
            target_date.strftime("%d/%m/%Y"),
            str(int(target_date.strftime("%Y%m%d"))),
            target_date.strftime("%Y/%m/%d")
        ]
        
        # Check each row (skip header row 1)
        for row_idx, cell_value in enumerate(date_col[1:], start=2):
            if not cell_value:
                continue
            
            # Normalize cell value
            cell_str = str(cell_value).strip()
            
            # Try parsing and comparing
            for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y%m%d", "%Y/%m/%d"]:
                try:
                    cell_date = datetime.strptime(cell_str, fmt)
                    if cell_date.date() == target_date.date():
                        return row_idx
                except ValueError:
                    continue
            
            # Direct string comparison
            if cell_str in target_date_strs:
                return row_idx
        
        return None
    
    def create_row_for_date(self, target_date: datetime) -> Optional[int]:
        """Create a new row for the given date in the target (Google Sheets or Excel)."""
        if self.test_mode:
            return self.find_or_create_row_in_excel(target_date)
        
        # Google Sheets implementation (existing code)
        try:
            # Get the date column index
            date_col_index = self.get_date_column_index()
            if not date_col_index:
                print(f"  Error: Could not find date column '{self.config['google_sheet'].get('date_column', 'Date')}'")
                print(f"  Available headers: {self._retry_gspread(self.worksheet.row_values, 1)[:10]}...")  # Show first 10 headers
                return None
            
            # Get all existing rows to find where to insert
            all_values = self._retry_gspread(self.worksheet.get_all_values)
            if not all_values:
                # Empty sheet, add header and first row
                headers = self._retry_gspread(self.worksheet.row_values, 1)
                if not headers:
                    return None
                new_row_num = 2
            else:
                # Find the right position to insert (keep dates sorted)
                date_col_values = self._retry_gspread(self.worksheet.col_values, date_col_index)
                new_row_num = len(date_col_values) + 1
                
                # Try to insert in chronological order
                for row_idx, cell_value in enumerate(date_col_values[1:], start=2):
                    if not cell_value:
                        new_row_num = row_idx
                        break
                    try:
                        # Try to parse the date
                        cell_str = str(cell_value).strip()
                        for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y%m%d", "%Y/%m/%d"]:
                            try:
                                cell_date = datetime.strptime(cell_str, fmt)
                                if cell_date.date() > target_date.date():
                                    new_row_num = row_idx
                                    break
                            except ValueError:
                                continue
                        if new_row_num < len(date_col_values) + 1:
                            break
                    except:
                        continue
            
            # Determine template row (row 2 should always be the template with formulas)
            template_row = 2
            
            # Insert a new row
            self._retry_gspread(self.worksheet.insert_row, [], new_row_num)
            
            # After insertion, the template row might have moved
            # If we inserted at row 2, template is now at row 3
            # If we inserted after row 2, template is still at row 2
            actual_template_row = template_row if new_row_num > template_row else template_row + 1
            
            # Copy formulas from template row to the new row
            # This ensures all formulas are preserved, even for the first data row
            headers = self._retry_gspread(self.worksheet.row_values, 1)
            num_cols = len(headers)
            
            # Copy formulas from template row to new row using batch update
            import string
            formula_updates = []
            
            for col_idx in range(1, num_cols + 1):
                # Skip the date column - we'll set that manually
                if col_idx == date_col_index:
                    continue
                
                # Convert to A1 notation
                col_letter = ''
                col_num = col_idx
                while col_num > 0:
                    col_num -= 1
                    col_letter = string.ascii_uppercase[col_num % 26] + col_letter
                    col_num //= 26
                
                source_cell = f"{col_letter}{actual_template_row}"
                target_cell = f"{col_letter}{new_row_num}"
                
                # Get the cell value with formula
                try:
                    cell = self._retry_gspread(self.worksheet.acell, source_cell, value_render_option='FORMULA')
                    if cell.value and str(cell.value).strip().startswith('='):
                        # Adjust formula references to point to the new row
                        formula = str(cell.value)
                        # Replace the template row number with the new row number in the formula
                        # This handles formulas like =IF(A2="","",TEXT(A2,"ddd")) -> =IF(A3="","",TEXT(A3,"ddd"))
                        adjusted_formula = formula.replace(f"A{actual_template_row}", f"A{new_row_num}")
                        adjusted_formula = adjusted_formula.replace(f"$A{actual_template_row}", f"$A{new_row_num}")
                        # Also replace the row number in other column references
                        import re
                        # Replace row numbers in cell references (e.g., B2, C2, etc.)
                        pattern = r'([A-Z]+\$?)(\d+)'
                        def replace_row(match):
                            col_ref = match.group(1)
                            row_num = int(match.group(2))
                            if row_num == actual_template_row:
                                return f"{col_ref}{new_row_num}"
                            return match.group(0)
                        adjusted_formula = re.sub(pattern, replace_row, adjusted_formula)
                        
                        formula_updates.append({
                            'range': target_cell,
                            'values': [[adjusted_formula]]
                        })
                except Exception as e:
                    # If we can't get the formula, skip this column
                    pass
            
            # Batch update all formulas at once
            # Use value_input_option='USER_ENTERED' to ensure formulas are interpreted as formulas, not text
            if formula_updates:
                for update in formula_updates:
                    try:
                        self._retry_gspread(
                            self.worksheet.update,
                            range_name=update['range'],
                            values=update['values'],
                            value_input_option='USER_ENTERED'  # This ensures formulas are treated as formulas, not text
                        )
                    except Exception as e:
                        print(f"    Warning: Could not copy formula to {update['range']}: {e}")
                print(f"  {CHECKMARK} Copied {len(formula_updates)} formulas from template row {actual_template_row} to row {new_row_num}")
            
            # Set the date in the date column using direct range update (more reliable)
            date_formatted = self.format_date_for_sheet(target_date)
            # Convert column index to A1 notation (e.g., 1 -> A, 2 -> B, 27 -> AA)
            import string
            col_letter = ''
            col_num = date_col_index
            while col_num > 0:
                col_num -= 1
                col_letter = string.ascii_uppercase[col_num % 26] + col_letter
                col_num //= 26
            date_range = f"{col_letter}{new_row_num}"
            
            # Always update the date column (even if it has a formula, we want to set the actual date)
            self._retry_gspread(
                self.worksheet.update,
                range_name=date_range,
                values=[[date_formatted]],
                value_input_option='USER_ENTERED'
            )
            
            print(f"  {CHECKMARK} Set date '{date_formatted}' in column {date_col_index} (row {new_row_num})")
            
            return new_row_num
            
        except Exception as e:
            print(f"  Error creating row: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def map_csv_to_sheet_columns(self, csv_file: Path, target_date: datetime) -> Dict[str, any]:
        """Map CSV data to Google Sheet columns based on csv_structure.json."""
        csv_filename = csv_file.name
        
        # Find matching mapping configuration from csv_structure.json
        mapping_config = self.csv_structure.get(csv_filename)
        
        if not mapping_config:
            print(f"  Warning: No mapping configuration found for {csv_filename}")
            return {}
        
        try:
            df = self._read_csv_safe(csv_file)
            if df is None or df.empty:
                return {}
            mapped_data = {}
            
            # Check for special mappings first (category pivot, etc.)
            special_mappings = mapping_config.get('special_mappings', {})
            if special_mappings:
                # Apply special mappings (these will generate multiple columns)
                # When special mappings exist, they replace regular column mappings
                for rule_name, rule_config in special_mappings.items():
                    values = self.apply_special_mapping(df, rule_config)
                    mapped_data.update(values)
            else:
                # Process regular column mappings (only if no special mappings)
                column_mappings = mapping_config.get('column_mappings', {})
                for csv_col, sheet_col in column_mappings.items():
                    if csv_col not in df.columns:
                        continue
                    
                    # Extract value (no rules in csv_structure.json, use default first value)
                    value = self.extract_value_from_csv(df, csv_col, None)
                    
                    if value is not None:
                        mapped_data[sheet_col] = value
            
            return mapped_data
            
        except Exception as e:
            print(f"  Error mapping {csv_filename}: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    def format_date_for_sheet(self, date: datetime) -> str:
        """Format date for Google Sheet (use format that matches existing sheet format)."""
        # Try to match existing format in sheet, default to YYYY-MM-DD
        return date.strftime("%Y-%m-%d")
    
    def extract_value_from_csv(self, df: pd.DataFrame, column: str, rule: Optional[Dict]) -> any:
        """Extract value from CSV column based on rules."""
        if column not in df.columns:
            return None
        
        if not rule:
            # Default: get first value
            return df[column].iloc[0]
        
        # Apply rule-based extraction
        rule_type = rule.get('type', 'first')
        
        if rule_type == 'first':
            return df[column].iloc[0]
        elif rule_type == 'total':
            # Find row with 'Total' in any column
            total_row = df[df.apply(lambda row: any('Total' in str(val) for val in row), axis=1)]
            if not total_row.empty:
                return total_row[column].iloc[0]
        elif rule_type == 'filter':
            # Filter by condition
            filter_col = rule.get('filter_column')
            filter_value = rule.get('filter_value')
            if filter_col and filter_col in df.columns:
                filtered = df[df[filter_col] == filter_value]
                if not filtered.empty:
                    return filtered[column].iloc[0]
        elif rule_type == 'sum':
            return df[column].sum()
        elif rule_type == 'category':
            # Extract value for specific category
            category_col = rule.get('category_column', 'Sales category')
            category_value = rule.get('category_value')
            if category_col in df.columns:
                category_row = df[df[category_col] == category_value]
                if not category_row.empty:
                    return category_row[column].iloc[0]
        
        return df[column].iloc[0]  # Fallback
    
    def apply_special_mapping(self, df: pd.DataFrame, rule_config: Dict) -> Dict[str, any]:
        """Apply special mapping rules (e.g., category breakdowns, category pivot)."""
        results = {}
        
        mapping_type = rule_config.get('type')
        
        if mapping_type == 'category_breakdown':
            category_col = rule_config.get('category_column', 'Sales category')
            value_col = rule_config.get('value_column', 'Net sales')
            target_columns = rule_config.get('target_columns', {})
            
            if category_col in df.columns and value_col in df.columns:
                for category, sheet_col in target_columns.items():
                    category_row = df[df[category_col] == category]
                    if not category_row.empty:
                        results[sheet_col] = category_row[value_col].iloc[0]
        
        elif mapping_type == 'category_pivot':
            # Transform category rows into columns (e.g., Food Items, Food Net sales, etc.)
            category_col = rule_config.get('category_column', 'Sales category')
            metrics = rule_config.get('metrics', [])
            categories = rule_config.get('categories', [])
            column_format = rule_config.get('column_name_format', '{category} {metric}')
            
            if category_col not in df.columns:
                return results
            
            # For each category and metric combination, create a column
            for category in categories:
                # Find the row for this category (case-insensitive and strip whitespace)
                category_clean = str(category).strip()
                # Try exact match first
                category_row = df[df[category_col].astype(str).str.strip() == category_clean]
                
                # If no exact match, try case-insensitive
                if category_row.empty:
                    category_row = df[df[category_col].astype(str).str.strip().str.lower() == category_clean.lower()]
                
                # If still no match, try partial match (e.g., "Non-Grat Svc Ch" might match "Non-Gratuity Service Charges")
                if category_row.empty:
                    # Extract key words from category (remove punctuation, split on spaces/hyphens)
                    category_words = category_clean.lower().replace('-', ' ').replace(',', ' ').split()
                    if len(category_words) > 0:
                        # Try to find rows where the category contains any of the key words
                        mask = df[category_col].astype(str).str.lower()
                        for word in category_words:
                            if len(word) > 2:  # Only use words longer than 2 characters
                                matches = mask.str.contains(word, na=False, regex=False)
                                if matches.any():
                                    category_row = df[matches]
                                    break
                
                if not category_row.empty:
                    for metric in metrics:
                        if metric in df.columns:
                            # Format column name (e.g., "Food Items", "Food Net sales")
                            column_name = column_format.format(category=category, metric=metric)
                            value = category_row[metric].iloc[0]
                            
                            # Handle NaN values and convert to appropriate type
                            if pd.notna(value):
                                # Convert numeric values appropriately
                                try:
                                    if isinstance(value, (int, float)):
                                        results[column_name] = value
                                    else:
                                        results[column_name] = value
                                except (ValueError, TypeError):
                                    results[column_name] = value
                else:
                    # Warn if category not found (for debugging)
                    print(f"    Warning: Category '{category}' not found in CSV for category pivot")
        
        elif mapping_type == 'category_pivot_combined':
            # Transform category rows into columns with combined primary and secondary category columns
            # (e.g., "Credit/debit - MASTERCARD Count", "Cash Amount", etc.)
            primary_col = rule_config.get('category_column_primary')
            secondary_col = rule_config.get('category_column_secondary')
            combine_format = rule_config.get('combine_format', '{primary} - {secondary}')
            metrics = rule_config.get('metrics', [])
            categories = rule_config.get('categories', [])
            column_format = rule_config.get('column_name_format', '{category} {metric}')
            
            if not primary_col or primary_col not in df.columns:
                return results
            
            # Create combined category column for matching
            def combine_categories(row):
                primary_val = str(row[primary_col]).strip() if pd.notna(row[primary_col]) else ''
                secondary_val = str(row[secondary_col]).strip() if secondary_col and secondary_col in df.columns and pd.notna(row[secondary_col]) else ''
                
                if secondary_val and secondary_val != '' and secondary_val.lower() != 'nan':
                    return combine_format.format(primary=primary_val, secondary=secondary_val)
                else:
                    return primary_val
            
            # Add combined category column to dataframe for matching
            df['_combined_category'] = df.apply(combine_categories, axis=1)
            
            # For each category and metric combination, create a column
            for category in categories:
                category_clean = str(category).strip()
                
                # Find the row for this combined category
                category_row = df[df['_combined_category'].astype(str).str.strip() == category_clean]
                
                # If no exact match, try case-insensitive
                if category_row.empty:
                    category_row = df[df['_combined_category'].astype(str).str.strip().str.lower() == category_clean.lower()]
                
                # If still no match, try partial match
                if category_row.empty:
                    category_words = category_clean.lower().replace('-', ' ').replace('/', ' ').replace(',', ' ').split()
                    if len(category_words) > 0:
                        mask = df['_combined_category'].astype(str).str.lower()
                        for word in category_words:
                            if len(word) > 2:  # Only use words longer than 2 characters
                                matches = mask.str.contains(word, na=False, regex=False)
                                if matches.any():
                                    category_row = df[matches]
                                    break
                
                if not category_row.empty:
                    for metric in metrics:
                        if metric in df.columns:
                            # Format column name (e.g., "Cash Count", "Credit/debit - MASTERCARD Amount")
                            column_name = column_format.format(category=category, metric=metric)
                            value = category_row[metric].iloc[0]
                            
                            # Handle NaN values and convert to appropriate type
                            if pd.notna(value):
                                # Convert numeric values appropriately
                                try:
                                    if isinstance(value, (int, float)):
                                        results[column_name] = value
                                    else:
                                        results[column_name] = value
                                except (ValueError, TypeError):
                                    results[column_name] = value
                else:
                    # Warn if category not found (for debugging)
                    print(f"    Warning: Combined category '{category}' not found in CSV for category pivot")
        
        return results
    
    def check_existing_data(self, row_num: int) -> bool:
        """Check if row already has data (beyond just the date) in the current worksheet."""
        if not row_num:
            return False
        
        if self.test_mode:
            # For Excel, check if row has data in the current worksheet
            if not self.excel_worksheet:
                return False
            try:
                # Check if row exists and has data beyond date column
                if row_num > self.excel_worksheet.max_row:
                    return False
                
                # Check if any cells beyond the first column (Date) have data
                for col in range(2, self.excel_worksheet.max_column + 1):
                    cell_value = self.excel_worksheet.cell(row=row_num, column=col).value
                    if cell_value is not None and str(cell_value).strip():
                        return True
                return False
            except Exception as e:
                return False
        
        # Google Sheets implementation
        if not self.worksheet:
            return False
        row_values = self._retry_gspread(self.worksheet.row_values, row_num)
        # Check if any cells beyond the first few have data
        return len([v for v in row_values[1:] if v]) > 0
    
    def cell_has_formula(self, row_num: int, col_index: int) -> bool:
        """Check if a cell contains a formula."""
        try:
            # Convert column index to A1 notation
            import string
            col_letter = ''
            col_num = col_index
            while col_num > 0:
                col_num -= 1
                col_letter = string.ascii_uppercase[col_num % 26] + col_letter
                col_num //= 26
            
            cell_range = f"{col_letter}{row_num}"
            
            # Try to get the cell with formula rendering
            try:
                cell = self._retry_gspread(self.worksheet.acell, cell_range, value_render_option='FORMULA')
                # Check if the cell value starts with '=' (formula indicator)
                if cell.value and str(cell.value).strip().startswith('='):
                    return True
            except:
                # If that fails, try getting the cell normally and check if it's a formula
                try:
                    cell = self._retry_gspread(self.worksheet.acell, cell_range)
                    # In Google Sheets API, formulas are indicated differently
                    # We can also check the cell's formulaValue property if available
                    if hasattr(cell, 'formulaValue') and cell.formulaValue:
                        return True
                except:
                    pass
            
            return False
        except Exception as e:
            # If we can't check, assume no formula to be safe (we'll update it)
            return False
    
    def update_google_sheet(self, row_num: int, data: Dict[str, any]) -> bool:
        """Update Google Sheet row with mapped data. Preserves formulas in cells.
        
        IMPORTANT: Only updates columns that are in the 'data' dictionary, which only
        contains columns specified in the config.json column_mappings. Other columns
        are never touched by this function.
        """
        if not row_num:
            print("  Error: Cannot update - row number not found")
            return False
        
        try:
            # Get column headers
            headers = self._retry_gspread(self.worksheet.row_values, 1)
            
            # Sort data to put Date column first (same as test mode)
            date_col_name = self.config['google_sheet'].get('date_column', 'Date')
            sorted_data = {}
            if date_col_name in data:
                sorted_data[date_col_name] = data[date_col_name]
            for key, value in data.items():
                if key != date_col_name:
                    sorted_data[key] = value
            
            # Prepare update batch
            # NOTE: Only columns in 'data' (from column_mappings) will be updated
            # All other columns in the sheet are left untouched
            updates = []
            skipped_formulas = []
            for sheet_col, value in sorted_data.items():
                # Try exact match first, then case-insensitive
                try:
                    col_index = headers.index(sheet_col) + 1
                except ValueError:
                    # Try case-insensitive match
                    try:
                        col_index = next(i for i, h in enumerate(headers, 1) if h.lower() == sheet_col.lower())
                        print(f"  Note: Found column '{headers[col_index-1]}' for '{sheet_col}' (case-insensitive)")
                    except StopIteration:
                        print(f"  Warning: Column '{sheet_col}' not found in sheet headers")
                        continue
                
                # Check if cell has a formula - if so, skip updating it
                # Only check for formulas if the cell is not empty (to avoid false positives)
                try:
                    # Get current cell value to check if it exists
                    import string
                    check_col_letter = ''
                    check_col_num = col_index
                    while check_col_num > 0:
                        check_col_num -= 1
                        check_col_letter = string.ascii_uppercase[check_col_num % 26] + check_col_letter
                        check_col_num //= 26
                    check_range = f"{check_col_letter}{row_num}"
                    current_cell = self._retry_gspread(self.worksheet.acell, check_range, value_render_option='FORMULA')
                    
                    # If cell has a formula (starts with '='), skip it
                    if current_cell.value and str(current_cell.value).strip().startswith('='):
                        skipped_formulas.append(sheet_col)
                        continue
                except:
                    # If we can't check, proceed with update (safer to update than skip)
                    pass
                
                # Convert value to appropriate format
                if pd.isna(value):
                    cell_value = ""
                elif isinstance(value, (int, float)):
                    cell_value = float(value)
                else:
                    cell_value = str(value)
                
                # Convert column index to A1 notation (e.g., 1 -> A, 2 -> B, 27 -> AA)
                import string
                col_letter = ''
                col_num = col_index
                while col_num > 0:
                    col_num -= 1
                    col_letter = string.ascii_uppercase[col_num % 26] + col_letter
                    col_num //= 26
                
                updates.append({
                    'range': f"{col_letter}{row_num}",
                    'values': [[cell_value]]
                })
            
            # Batch update
            # Use value_input_option='USER_ENTERED' to ensure proper formatting
            if updates:
                for update in updates:
                    self._retry_gspread(
                        self.worksheet.update,
                        range_name=update['range'],
                        values=update['values'],
                        value_input_option='USER_ENTERED'
                    )
                
                print(f"  {CHECKMARK} Updated {len(updates)} columns in row {row_num}")
                if skipped_formulas:
                    print(f"  Note: Skipped {len(skipped_formulas)} columns with formulas: {', '.join(skipped_formulas[:3])}{'...' if len(skipped_formulas) > 3 else ''}")
                date_col_name = self.config['google_sheet'].get('date_column', 'Date')
                if date_col_name in data:
                    print(f"  {CHECKMARK} Date '{data[date_col_name]}' should be in row {row_num}")
                return True
            else:
                if skipped_formulas:
                    print(f"  Note: All columns had formulas, no data updated (formulas preserved)")
                else:
                    print("  Warning: No valid updates to perform")
                return False
                
        except Exception as e:
            print(f"  Error updating Google Sheet: {e}")
            return False
    
    def create_excel_backup(self, excel_path: Path) -> Optional[Path]:
        """Create a backup of the Excel file before updating."""
        try:
            if not excel_path.exists():
                return None
            
            # Don't backup backup files
            if "backup" in excel_path.name.lower():
                print(f"  {WARNING} Skipping backup creation for backup file")
                return None
            
            # Create backup filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = excel_path.parent / f"{excel_path.stem}_backup_{timestamp}{excel_path.suffix}"
            
            shutil.copy2(excel_path, backup_path)
            print(f"  {CHECKMARK} Created backup: {backup_path.name}")
            return backup_path
        except Exception as e:
            print(f"  {WARNING} Could not create backup: {e}")
            return None
    
    def load_excel_file(self) -> bool:
        """Load Excel file for test mode."""
        try:
            excel_path = Path(__file__).parent / self.excel_file
            
            # If file not found in root, search in subdirectories
            # Exclude backup files (files with "backup" in name)
            if not excel_path.exists():
                base_path = Path(__file__).parent
                found_path = None
                
                # Search in subdirectories
                for subdir in base_path.iterdir():
                    if subdir.is_dir():
                        potential_path = subdir / self.excel_file
                        # Skip backup files
                        if potential_path.exists() and "backup" not in potential_path.name.lower():
                            found_path = potential_path
                            break
                
                if found_path:
                    excel_path = found_path
                else:
                    print(f"  {CROSS} Excel file not found: {self.excel_file}")
                    print(f"     Searched in: {base_path} and subdirectories")
                    print(f"     Note: Backup files are excluded from search")
                    return False
            
            # Double-check: Don't use backup files
            if "backup" in excel_path.name.lower():
                print(f"  {CROSS} Found Excel file is a backup file: {excel_path.name}")
                print(f"     Please use the original file, not the backup")
                return False
            
            # Store the actual path found
            self.excel_file_path = excel_path
            
            # Create backup before loading
            self.create_excel_backup(excel_path)
            
            # Load workbook with openpyxl to preserve formatting
            self.workbook = load_workbook(excel_path)
            
            # Check if we should use test sheet for column creation
            if self.test_mode and self.auto_create_columns and self.test_sheet_name:
                # Use test sheet for dynamic column creation
                self.using_test_sheet = True
                if not self.create_or_get_test_sheet():
                    return False
                # Save workbook after creating test sheet so it's available for reading
                try:
                    self.workbook.save(excel_path)
                except PermissionError:
                    print(f"  {WARNING} Could not save Excel file (file may be open). Will try again when updating data.")
                print(f"  {CHECKMARK} Loaded Excel file: {excel_path.name}")
            else:
                # Use regular sheet
                if self.excel_sheet_name not in self.workbook.sheetnames:
                    print(f"  {CROSS} Sheet '{self.excel_sheet_name}' not found in Excel file")
                    print(f"     Available sheets: {', '.join(self.workbook.sheetnames)}")
                    return False
                
                self.excel_worksheet = self.workbook[self.excel_sheet_name]
                print(f"  {CHECKMARK} Loaded Excel file: {excel_path.name}")
                print(f"  {CHECKMARK} Using sheet: {self.excel_sheet_name}")
            
            return True
            
        except Exception as e:
            print(f"  {CROSS} Error loading Excel file: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def create_or_get_sheet(self, sheet_name: str) -> bool:
        """Create or get a sheet by name for dynamic column creation. Uses same header styles as Daily Ops."""
        try:
            from copy import copy
            
            if not self.workbook:
                print(f"  {CROSS} Workbook not loaded")
                return False
            
            # Check if sheet exists
            if sheet_name in self.workbook.sheetnames:
                self.excel_worksheet = self.workbook[sheet_name]
            else:
                # Create new empty sheet
                self.excel_worksheet = self.workbook.create_sheet(sheet_name)
                print(f"  {CHECKMARK} Created new empty sheet: {sheet_name}")
            
            # Get header style from original "Daily Ops" sheet to match formatting
            if self.excel_sheet_name in self.workbook.sheetnames:
                source_sheet = self.workbook[self.excel_sheet_name]
                if source_sheet.max_row > 0 and source_sheet.max_column > 0:
                    # Get style from first row, first column as template
                    template_cell = source_sheet.cell(row=1, column=1)
                    self.header_style = {
                        'font': copy(template_cell.font) if template_cell.font else Font(bold=True, size=11),
                        'fill': copy(template_cell.fill) if template_cell.fill else None,
                        'alignment': copy(template_cell.alignment) if template_cell.alignment else Alignment(horizontal='center', vertical='center'),
                        'border': copy(template_cell.border) if template_cell.border else None
                    }
                else:
                    # Default header style
                    self.header_style = {
                        'font': Font(bold=True, size=11),
                        'fill': None,
                        'alignment': Alignment(horizontal='center', vertical='center'),
                        'border': None
                    }
            else:
                # Default header style if source sheet not found
                self.header_style = {
                    'font': Font(bold=True, size=11),
                    'fill': None,
                    'alignment': Alignment(horizontal='center', vertical='center'),
                    'border': None
                }
            
            return True
        except Exception as e:
            print(f"  {CROSS} Error creating/getting sheet '{sheet_name}': {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def create_or_get_test_sheet(self) -> bool:
        """Create or get the test sheet for dynamic column creation."""
        return self.create_or_get_sheet(self.test_sheet_name)
    
    def ensure_column_exists(self, column_name: str, existing_headers: list = None) -> int:
        """Ensure a column exists in the current worksheet. Returns column index (1-based).
        
        Args:
            column_name: Name of the column to ensure exists
            existing_headers: Optional list of existing headers to avoid re-reading
        """
        if not self.excel_worksheet:
            return None
        
        try:
            from copy import copy
            
            # Get existing headers from first row if not provided
            if existing_headers is None:
                headers = []
                for col in range(1, self.excel_worksheet.max_column + 1):
                    cell_value = self.excel_worksheet.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value).strip())
                    else:
                        break
            else:
                headers = existing_headers[:]  # Copy to avoid modifying original
            
            # Check if column already exists (case-insensitive)
            for idx, header in enumerate(headers, start=1):
                if header.lower() == column_name.lower():
                    return idx
            
            # Column doesn't exist, add it
            date_col_name = self.config['google_sheet'].get('date_column', 'Date')
            
            # If this is the Date column, ensure it's first
            if column_name == date_col_name:
                # If headers exist but Date is not first, we need to insert it
                if headers and headers[0].lower() != date_col_name.lower():
                    # Date column should be first - this case is handled in update_excel_file
                    # by ensuring Date is created first before other columns
                    pass
                
                # Add Date column at position 1
                # If there are existing columns, we need to shift them right
                if self.excel_worksheet.max_column > 0:
                    # Shift all existing columns one position to the right
                    max_col = self.excel_worksheet.max_column
                    for row in range(1, self.excel_worksheet.max_row + 1):
                        # Start from the rightmost column and move left
                        for col in range(max_col, 0, -1):
                            source_cell = self.excel_worksheet.cell(row=row, column=col)
                            target_cell = self.excel_worksheet.cell(row=row, column=col + 1)
                            target_cell.value = source_cell.value
                            if source_cell.font:
                                target_cell.font = copy(source_cell.font)
                            if source_cell.fill:
                                target_cell.fill = copy(source_cell.fill)
                            if source_cell.alignment:
                                target_cell.alignment = copy(source_cell.alignment)
                            if source_cell.border:
                                target_cell.border = copy(source_cell.border)
                            target_cell.number_format = source_cell.number_format
                
                # Now add Date header at column 1
                date_cell = self.excel_worksheet.cell(row=1, column=1)
                date_cell.value = date_col_name
                
                # Apply header style
                if hasattr(self, 'header_style'):
                    if self.header_style.get('font'):
                        date_cell.font = self.header_style['font']
                    if self.header_style.get('fill'):
                        date_cell.fill = self.header_style['fill']
                    if self.header_style.get('alignment'):
                        date_cell.alignment = self.header_style['alignment']
                    if self.header_style.get('border'):
                        date_cell.border = self.header_style['border']
                
                print(f"  {CHECKMARK} Added Date column as first column")
                return 1
            else:
                # Regular column - add at the end
                # But first check if Date column exists, if not add it first
                date_exists = False
                for header in headers:
                    if header.lower() == date_col_name.lower():
                        date_exists = True
                        break
                
                if not date_exists:
                    # Date doesn't exist, add it first (recursively, but Date won't recurse)
                    self.ensure_column_exists(date_col_name, headers)
                    # Refresh headers count
                    new_col_index = self.excel_worksheet.max_column + 1
                else:
                    # Date exists, add new column after existing columns
                    new_col_index = self.excel_worksheet.max_column + 1
                    if new_col_index == 1:
                        # Only Date column exists, so new column goes to position 2
                        new_col_index = 2
                
                # Add new column header
                new_cell = self.excel_worksheet.cell(row=1, column=new_col_index)
                new_cell.value = column_name
                
                # Apply header style matching original sheet
                if hasattr(self, 'header_style'):
                    if self.header_style.get('font'):
                        new_cell.font = self.header_style['font']
                    if self.header_style.get('fill'):
                        new_cell.fill = self.header_style['fill']
                    if self.header_style.get('alignment'):
                        new_cell.alignment = self.header_style['alignment']
                    if self.header_style.get('border'):
                        new_cell.border = self.header_style['border']
                
                print(f"  {CHECKMARK} Created new column: {column_name}")
                return new_col_index
            
        except Exception as e:
            print(f"  {WARNING} Error ensuring column exists for '{column_name}': {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def find_or_create_row_in_excel(self, target_date: datetime) -> Optional[int]:
        """Find or create a row for the target date in Excel file."""
        if not self.excel_worksheet:
            return None
        
        try:
            date_column = self.config['google_sheet'].get('date_column', 'Date')
            target_date_str = target_date.strftime("%Y-%m-%d")
            
            # For sheets with auto-create enabled, handle empty sheet case
            if self.auto_create_columns:
                # Check if sheet is empty (no headers or no data rows)
                if self.excel_worksheet.max_row == 0 or (self.excel_worksheet.max_row == 1 and not self.excel_worksheet.cell(row=1, column=1).value):
                    # Empty sheet, create first row with Date column
                    date_col_index = 1
                    date_cell = self.excel_worksheet.cell(row=1, column=date_col_index)
                    date_cell.value = date_column
                    # Apply header style
                    if hasattr(self, 'header_style'):
                        if self.header_style.get('font'):
                            date_cell.font = self.header_style['font']
                        if self.header_style.get('fill'):
                            date_cell.fill = self.header_style['fill']
                        if self.header_style.get('alignment'):
                            date_cell.alignment = self.header_style['alignment']
                        if self.header_style.get('border'):
                            date_cell.border = self.header_style['border']
                
                # Now read the sheet to find existing rows
                if not self.excel_file_path:
                    print(f"  {CROSS} Excel file path not set")
                    return None
                
                # For sheets with auto-create, we can work directly with the worksheet object
                # Check if sheet is empty by checking if there are any headers
                has_headers = False
                if self.excel_worksheet.max_row >= 1:
                    for col in range(1, self.excel_worksheet.max_column + 1):
                        if self.excel_worksheet.cell(row=1, column=col).value:
                            has_headers = True
                            break
                
                if self.excel_worksheet.max_row == 0 or not has_headers:
                    # Empty sheet - treat as empty DataFrame
                    df = pd.DataFrame()
                else:
                    # Try to read from file, but if it fails (file open or sheet not saved), work with worksheet directly
                    try:
                        excel_path = self.excel_file_path
                        active_sheet_name = self.test_sheet_name
                        # Try to save first
                        try:
                            self.workbook.save(excel_path)
                        except PermissionError:
                            pass  # File might be open, continue anyway
                        
                        df = pd.read_excel(excel_path, sheet_name=active_sheet_name, header=0)
                    except (ValueError, PermissionError, FileNotFoundError):
                        # Can't read from file, build DataFrame from worksheet directly
                        # Read headers from first row
                        headers = []
                        for col in range(1, self.excel_worksheet.max_column + 1):
                            cell_value = self.excel_worksheet.cell(row=1, column=col).value
                            if cell_value:
                                headers.append(str(cell_value).strip())
                            else:
                                break
                        
                        # Read data rows
                        data_rows = []
                        for row in range(2, self.excel_worksheet.max_row + 1):
                            row_data = {}
                            for col_idx, header in enumerate(headers, start=1):
                                cell_value = self.excel_worksheet.cell(row=row, column=col_idx).value
                                row_data[header] = cell_value
                            if any(v is not None for v in row_data.values()):  # Only add non-empty rows
                                data_rows.append(row_data)
                        
                        df = pd.DataFrame(data_rows) if data_rows else pd.DataFrame(columns=headers)
                
                # Check if Date column exists
                if len(df.columns) == 0 or date_column not in df.columns:
                    # No Date column yet or empty sheet, create new row
                    new_row_num = 2  # Row 1 is header, row 2 is first data row
                    print(f"  {CHECKMARK} Creating new row {new_row_num} for date {target_date_str}")
                    
                    # Ensure Date column exists
                    date_col_index = 1
                    if self.excel_worksheet.max_column == 0 or self.excel_worksheet.cell(row=1, column=1).value != date_column:
                        date_cell = self.excel_worksheet.cell(row=1, column=date_col_index)
                        date_cell.value = date_column
                        # Apply header style
                        if hasattr(self, 'header_style'):
                            if self.header_style.get('font'):
                                date_cell.font = self.header_style['font']
                            if self.header_style.get('fill'):
                                date_cell.fill = self.header_style['fill']
                            if self.header_style.get('alignment'):
                                date_cell.alignment = self.header_style['alignment']
                            if self.header_style.get('border'):
                                date_cell.border = self.header_style['border']
                    
                    # Set date value
                    date_data_cell = self.excel_worksheet.cell(row=new_row_num, column=date_col_index)
                    date_data_cell.value = target_date
                    date_data_cell.number_format = 'yyyy-mm-dd'
                    
                    return new_row_num
                
                # Try to find existing row with this date
                for idx, row_date in enumerate(df[date_column], start=2):
                    if pd.notna(row_date):
                        # Convert Excel date to string for comparison
                        if isinstance(row_date, pd.Timestamp):
                            row_date_str = row_date.strftime("%Y-%m-%d")
                        elif isinstance(row_date, datetime):
                            row_date_str = row_date.strftime("%Y-%m-%d")
                        elif isinstance(row_date, str):
                            try:
                                row_date_str = pd.to_datetime(row_date).strftime("%Y-%m-%d")
                            except:
                                row_date_str = str(row_date).strip()
                        else:
                            row_date_str = str(row_date).strip()
                        
                        if row_date_str == target_date_str:
                            print(f"  {CHECKMARK} Found existing row {idx} for date {target_date_str}")
                            return idx
                
                # No existing row found, create new one at the end
                new_row_num = len(df) + 2  # +1 for header, +1 for new row
                print(f"  {CHECKMARK} Creating new row {new_row_num} for date {target_date_str}")
                
                # Set the date in the date column (always column 1 for test sheet)
                date_col_index = 1
                date_cell = self.excel_worksheet.cell(row=new_row_num, column=date_col_index)
                date_cell.value = target_date
                date_cell.number_format = 'yyyy-mm-dd'
                
                return new_row_num
            else:
                # Regular sheet - use existing logic
                if not self.excel_file_path:
                    print(f"  {CROSS} Excel file path not set")
                    return None
                
                excel_path = self.excel_file_path
                df = pd.read_excel(excel_path, sheet_name=self.excel_sheet_name, header=0)
                
                if date_column not in df.columns:
                    print(f"  {CROSS} Date column '{date_column}' not found in Excel sheet")
                    print(f"     Available columns: {', '.join(df.columns[:10])}...")
                    return None
                
                # Format target date for comparison
                target_date_str = target_date.strftime("%Y-%m-%d")
                
                # Try to find existing row with this date
                for idx, row_date in enumerate(df[date_column], start=2):
                    if pd.notna(row_date):
                        if isinstance(row_date, pd.Timestamp):
                            row_date_str = row_date.strftime("%Y-%m-%d")
                        elif isinstance(row_date, datetime):
                            row_date_str = row_date.strftime("%Y-%m-%d")
                        elif isinstance(row_date, str):
                            try:
                                row_date_str = pd.to_datetime(row_date).strftime("%Y-%m-%d")
                            except:
                                row_date_str = str(row_date).strip()
                        else:
                            row_date_str = str(row_date).strip()
                        
                        if row_date_str == target_date_str:
                            print(f"  {CHECKMARK} Found existing row {idx} for date {target_date_str}")
                            return idx
                
                # No existing row found, create new one at the end
                new_row_num = len(df) + 2
                print(f"  {CHECKMARK} Creating new row {new_row_num} for date {target_date_str}")
                
                date_col_index = list(df.columns).index(date_column) + 1
                date_cell = self.excel_worksheet.cell(row=new_row_num, column=date_col_index)
                date_cell.value = target_date
                date_cell.number_format = 'yyyy-mm-dd'
                
                return new_row_num
            
        except Exception as e:
            print(f"  {CROSS} Error finding/creating row in Excel: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def update_excel_file(self, row_num: int, data: Dict[str, any]) -> bool:
        """Update Excel file row with mapped data. Preserves formatting and formulas."""
        if not row_num or not self.excel_worksheet:
            print("  Error: Cannot update - row number or worksheet not found")
            return False
        
        try:
            # Get headers from the first row of the worksheet
            headers = []
            for col in range(1, self.excel_worksheet.max_column + 1):
                cell_value = self.excel_worksheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
                else:
                    break
            
            updates_count = 0
            skipped_formulas = []
            
            # With auto-create enabled, ensure all columns exist first
            if self.auto_create_columns:
                # Ensure Date column exists first
                date_col_name = self.config['google_sheet'].get('date_column', 'Date')
                if date_col_name in data:
                    self.ensure_column_exists(date_col_name)
                
                # Ensure all other mapped columns exist
                for sheet_col in data.keys():
                    if sheet_col != date_col_name:  # Date already handled
                        self.ensure_column_exists(sheet_col)
                
                # Refresh headers after creating columns
                headers = []
                for col in range(1, self.excel_worksheet.max_column + 1):
                    cell_value = self.excel_worksheet.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value).strip())
                    else:
                        break
            
            # Sort data to put Date column first
            date_col_name = self.config['google_sheet'].get('date_column', 'Date')
            sorted_data = {}
            if date_col_name in data:
                sorted_data[date_col_name] = data[date_col_name]
            for key, value in data.items():
                if key != date_col_name:
                    sorted_data[key] = value
            
            for sheet_col, value in sorted_data.items():
                # Find or create column index
                col_index = None
                
                # Try exact match first
                try:
                    col_index = headers.index(sheet_col) + 1  # +1 for Excel 1-based indexing
                except ValueError:
                    # Try case-insensitive match
                    try:
                        col_index = next(i for i, h in enumerate(headers, 1) if h.lower() == sheet_col.lower())
                    except StopIteration:
                        # Column not found
                        if self.auto_create_columns:
                            # Auto-create the column
                            col_index = self.ensure_column_exists(sheet_col, headers)
                            if col_index:
                                # Refresh headers after creating column
                                headers = []
                                for col in range(1, self.excel_worksheet.max_column + 1):
                                    cell_value = self.excel_worksheet.cell(row=1, column=col).value
                                    if cell_value:
                                        headers.append(str(cell_value).strip())
                                    else:
                                        break
                        else:
                            print(f"  Warning: Column '{sheet_col}' not found in Excel headers")
                            continue
                
                if not col_index:
                    continue
                
                # Get the cell
                cell = self.excel_worksheet.cell(row=row_num, column=col_index)
                
                # Check if cell has a formula - preserve it if so
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.strip().startswith('='):
                        skipped_formulas.append(sheet_col)
                        continue
                    # Also check if cell has a formula in its data_type property
                    if hasattr(cell, 'data_type') and cell.data_type == 'f':
                        skipped_formulas.append(sheet_col)
                        continue
                
                # Save current formatting (avoid deprecation warnings)
                from copy import copy
                current_font = copy(cell.font) if cell.font else None
                current_fill = copy(cell.fill) if cell.fill else None
                current_alignment = copy(cell.alignment) if cell.alignment else None
                current_border = copy(cell.border) if cell.border else None
                current_number_format = cell.number_format
                
                # Update value
                if pd.isna(value):
                    cell.value = None
                elif isinstance(value, (int, float)):
                    cell.value = float(value)
                else:
                    cell.value = str(value)
                
                # Restore formatting
                if current_font:
                    cell.font = current_font
                if current_fill:
                    cell.fill = current_fill
                if current_alignment:
                    cell.alignment = current_alignment
                if current_border:
                    cell.border = current_border
                cell.number_format = current_number_format
                
                updates_count += 1
            
            # Save the workbook
            if updates_count > 0 or len(skipped_formulas) > 0:
                if not self.excel_file_path:
                    print(f"  {CROSS} Excel file path not set, cannot save")
                    return False
                
                try:
                    self.workbook.save(self.excel_file_path)
                    print(f"  {CHECKMARK} Updated {updates_count} columns in Excel row {row_num}")
                    if skipped_formulas:
                        print(f"  Note: Skipped {len(skipped_formulas)} columns with formulas: {', '.join(skipped_formulas[:3])}{'...' if len(skipped_formulas) > 3 else ''}")
                    return True
                except PermissionError:
                    print(f"  {WARNING} Could not save Excel file (file may be open in another program)")
                    print(f"  {WARNING} Please close the Excel file and run the script again to save changes")
                    print(f"  Note: {updates_count} columns were prepared for update but not saved")
                    return False
            else:
                print("  Warning: No valid updates to perform")
                return False
                
        except Exception as e:
            print(f"  Error updating Excel file: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def update_target(self, row_num: int, data: Dict[str, any], target_date: datetime) -> bool:
        """Route to either Google Sheets or Excel based on test_mode."""
        if self.test_mode:
            return self.update_excel_file(row_num, data)
        else:
            return self.update_google_sheet(row_num, data)
    
    def check_week_ending_exists(self, tab_name: str, week_ending_date: datetime) -> Tuple[bool, int]:
        """Check if Date exists in the tab and return (exists, row_count)."""
        expected_headers = {'date', 'week_ending_date', 'week ending date'}
        if self.test_mode:
            # Excel version
            if not self.workbook or tab_name not in self.workbook.sheetnames:
                return False, 0
            
            worksheet = self.workbook[tab_name]
            week_ending_str = week_ending_date.strftime("%Y-%m-%d")
            row_count = 0
            
            # Find Date column (should be column A / column 1)
            if worksheet.max_row == 0:
                return False, 0
            
            # Check header row for Date column
            header_cell = worksheet.cell(row=1, column=1).value
            if header_cell and str(header_cell).strip().lower() not in expected_headers:
                # If header doesn't match, try to find it
                for col in range(1, worksheet.max_column + 1):
                    header_val = worksheet.cell(row=1, column=col).value
                    if header_val and str(header_val).strip().lower() in expected_headers:
                        # Found it, but it should be column 1 - continue anyway
                        break
            
            # Check all data rows (starting from row 2)
            for row_idx in range(2, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row_idx, column=1).value
                if cell_value is not None:
                    # Convert to string for comparison
                    if isinstance(cell_value, datetime):
                        cell_str = cell_value.strftime("%Y-%m-%d")
                    elif isinstance(cell_value, pd.Timestamp):
                        cell_str = cell_value.strftime("%Y-%m-%d")
                    else:
                        try:
                            cell_str = pd.to_datetime(str(cell_value)).strftime("%Y-%m-%d")
                        except:
                            cell_str = str(cell_value).strip()
                    
                    if cell_str == week_ending_str:
                        row_count += 1
            
            return row_count > 0, row_count
        else:
            # Google Sheets version
            if not self.sheet:
                return False, 0
            
            try:
                worksheet = self._get_worksheet(tab_name)
            except gspread.exceptions.WorksheetNotFound:
                return False, 0
            
            week_ending_str = week_ending_date.strftime("%Y-%m-%d")
            row_count = 0
            
            # Get all values from the first column (Date column)
            try:
                all_values = self._retry_gspread(worksheet.col_values, 1)  # Column A (index 1)
                if len(all_values) <= 1:  # Only header or empty
                    return False, 0
                
                # Check all data rows (starting from row 2, index 1)
                for cell_value in all_values[1:]:  # Skip header
                    if cell_value:
                        # Convert to string for comparison
                        try:
                            cell_str = pd.to_datetime(str(cell_value)).strftime("%Y-%m-%d")
                        except:
                            cell_str = str(cell_value).strip()
                        
                        if cell_str == week_ending_str:
                            row_count += 1
            except Exception as e:
                return False, 0
            
            return row_count > 0, row_count
    
    def ask_user_override(self, tab_name: str, week_ending_date: datetime, row_count: int) -> bool:
        """Ask user if they want to override existing data."""
        date_str = week_ending_date.strftime("%Y-%m-%d")
        print(f"\n      {WARNING} Date {date_str} already exists")
        print(f"      Found {row_count} existing row(s) with this date")
        
        while True:
            response = input(f"      Override {row_count} existing row(s)? (yes/no): ").strip().lower()
            if response in ['yes', 'y']:
                return True
            elif response in ['no', 'n']:
                return False
            else:
                print(f"      Please enter 'yes' or 'no'")
    
    def delete_rows_with_week_ending(self, tab_name: str, week_ending_date: datetime) -> int:
        """Delete all rows with the given date. Returns number of rows deleted."""
        if self.test_mode:
            # Excel version
            if not self.workbook or tab_name not in self.workbook.sheetnames:
                return 0
            
            worksheet = self.workbook[tab_name]
            week_ending_str = week_ending_date.strftime("%Y-%m-%d")
            rows_to_delete = []
            
            # Find rows to delete (starting from row 2, skipping header)
            for row_idx in range(2, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row_idx, column=1).value
                if cell_value is not None:
                    # Convert to string for comparison
                    if isinstance(cell_value, datetime):
                        cell_str = cell_value.strftime("%Y-%m-%d")
                    elif isinstance(cell_value, pd.Timestamp):
                        cell_str = cell_value.strftime("%Y-%m-%d")
                    else:
                        try:
                            cell_str = pd.to_datetime(str(cell_value)).strftime("%Y-%m-%d")
                        except:
                            cell_str = str(cell_value).strip()
                    
                    if cell_str == week_ending_str:
                        rows_to_delete.append(row_idx)
            
            # Delete rows in reverse order (from bottom to top) to avoid index shifting issues
            deleted_count = 0
            for row_idx in reversed(rows_to_delete):
                worksheet.delete_rows(row_idx)
                deleted_count += 1
            
            return deleted_count
        else:
            # Google Sheets version
            if not self.sheet:
                return 0
            
            try:
                worksheet = self._get_worksheet(tab_name)
            except gspread.exceptions.WorksheetNotFound:
                return 0
            
            week_ending_str = week_ending_date.strftime("%Y-%m-%d")
            rows_to_delete = []
            
            # Get all values from the first column (Date column)
            try:
                all_values = self._retry_gspread(worksheet.col_values, 1)  # Column A (index 1)
                if len(all_values) <= 1:  # Only header or empty
                    return 0
                
                # Find rows to delete (starting from row 2, index 1 in list but row 2 in sheet)
                for idx, cell_value in enumerate(all_values[1:], start=2):  # Start from row 2
                    if cell_value:
                        # Convert to string for comparison
                        try:
                            cell_str = pd.to_datetime(str(cell_value)).strftime("%Y-%m-%d")
                        except:
                            cell_str = str(cell_value).strip()
                        
                        if cell_str == week_ending_str:
                            rows_to_delete.append(idx)
            except Exception as e:
                return 0
            
            # Delete rows in reverse order (from bottom to top) to avoid index shifting issues
            deleted_count = 0
            for row_idx in reversed(rows_to_delete):
                try:
                    worksheet.delete_rows(row_idx)
                    deleted_count += 1
                except Exception as e:
                    continue  # Continue deleting other rows even if one fails
            
            return deleted_count
    
    def _column_index_to_a1(self, col_idx: int) -> str:
        """Convert column index (1-based) to A1 notation (e.g., 1 -> A, 2 -> B, 27 -> AA)."""
        import string
        col_letter = ''
        col_num = col_idx
        while col_num > 0:
            col_num -= 1
            col_letter = string.ascii_uppercase[col_num % 26] + col_letter
            col_num //= 26
        return col_letter

    def _filter_total_rows(self, df: pd.DataFrame, csv_file: Path) -> pd.DataFrame:
        """Remove rows where the first column contains a Total marker."""
        if df.empty:
            return df
        first_col = df.columns[0]
        total_markers = {"total", "totals"}
        series = df[first_col].astype(str).str.strip().str.lower()
        mask = series.isin(total_markers)
        removed = int(mask.sum())
        if removed:
            print(f"      {WARNING} Skipped {removed} Total row(s) in {csv_file.name}")
        return df[~mask].copy()
    
    def append_csv_to_excel_tab(self, csv_file: Path, tab_name: str, week_ending_date: datetime) -> bool:
        """Append CSV data directly to Excel tab or Google Sheet, matching headers and adding Date column."""
        expected_headers = {'date', 'week_ending_date', 'week ending date'}
        if self.test_mode:
            # Excel version
            try:
                # Check if tab exists
                if not self.workbook:
                    print(f"  {CROSS} Workbook not loaded")
                    return False
                
                if tab_name not in self.workbook.sheetnames:
                    print(f"  {CROSS} Tab '{tab_name}' does not exist in Excel file")
                    print(f"     Available tabs: {', '.join(self.workbook.sheetnames)}")
                    return False
                
                worksheet = self.workbook[tab_name]
                
                # Read CSV file
                df = self._read_csv_safe(csv_file)
                if df is None or df.empty:
                    return False
                df = self._filter_total_rows(df, csv_file)
                
                # Get Excel headers (row 1)
                excel_headers = []
                for col in range(1, worksheet.max_column + 1):
                    header_cell = worksheet.cell(row=1, column=col).value
                    if header_cell:
                        excel_headers.append(str(header_cell).strip())
                    else:
                        break
                
                if not excel_headers:
                    print(f"  {CROSS} No headers found in tab '{tab_name}'")
                    return False
                
                # Date should be the first column
                if excel_headers[0].lower() not in expected_headers:
                    print(f"  {WARNING} First column should be 'Date', found: {excel_headers[0]}")
                    # Continue anyway
                
                # Map CSV column names to Excel column names (case-insensitive matching)
                # Create a mapping dictionary
                # Exclude Clasification column - it will use a formula instead
                csv_to_excel_mapping = {}
                unmapped_csv_cols = []
                
                for csv_col in df.columns:
                    csv_col_stripped = str(csv_col).strip()
                    # Skip Clasification column - it will use a formula
                    if csv_col_stripped.lower() in ['clasification', 'classification']:
                        continue
                    
                    # Try exact match first
                    if csv_col_stripped in excel_headers:
                        csv_to_excel_mapping[csv_col] = csv_col_stripped
                    else:
                        # Try case-insensitive match
                        found = False
                        for excel_col in excel_headers:
                            # Skip Clasification column
                            if excel_col.lower() in ['clasification', 'classification']:
                                continue
                            if csv_col_stripped.lower() == excel_col.lower():
                                csv_to_excel_mapping[csv_col] = excel_col
                                found = True
                                break
                        if not found:
                            unmapped_csv_cols.append(csv_col)
                
                # Warn about unmapped columns
                if unmapped_csv_cols:
                    print(f"      {WARNING} CSV columns not found in Excel (skipped): {', '.join(unmapped_csv_cols[:5])}{'...' if len(unmapped_csv_cols) > 5 else ''}")
                
                # Add Date column to dataframe (first column)
                # Use date only (no time) for the input date
                date_only = week_ending_date.date() if hasattr(week_ending_date, 'date') else week_ending_date
                df.insert(0, self.date_column_name, date_only)
                csv_to_excel_mapping[self.date_column_name] = excel_headers[0]  # Map to first Excel column
                
                # Determine starting row (after existing data)
                start_row = worksheet.max_row + 1
                if worksheet.max_row == 0:
                    start_row = 2  # Row 1 is header, data starts at row 2
                elif worksheet.max_row == 1:
                    # Only header row exists
                    start_row = 2
                else:
                    start_row = worksheet.max_row + 1
                
                # Append each row from CSV
                rows_appended = 0
                for idx, csv_row in df.iterrows():
                    row_num = start_row + rows_appended
                    
                    # Write data to Excel row, matching columns
                    for csv_col, excel_col in csv_to_excel_mapping.items():
                        # Find Excel column index
                        try:
                            col_idx = excel_headers.index(excel_col) + 1  # Excel is 1-based
                        except ValueError:
                            continue  # Skip if column not found
                        
                        # Get value from CSV
                        if csv_col == self.date_column_name:
                            # Use date only (no time) for the input date
                            value = week_ending_date.date() if hasattr(week_ending_date, 'date') else week_ending_date
                        else:
                            value = csv_row[csv_col]
                        
                        # Write to cell
                        cell = worksheet.cell(row=row_num, column=col_idx)
                        if pd.isna(value):
                            cell.value = None
                        elif csv_col == self.date_column_name:
                            # For date column, use date object directly (no time)
                            from datetime import date as date_type
                            if isinstance(value, datetime):
                                cell.value = value.date()
                            elif isinstance(value, date_type):
                                cell.value = value
                            else:
                                # Try to convert to date
                                try:
                                    cell.value = pd.to_datetime(value).date()
                                except:
                                    cell.value = value
                            cell.number_format = 'yyyy-mm-dd'
                        elif isinstance(value, (int, float)):
                            cell.value = float(value)
                        else:
                            cell.value = str(value)
                    
                    rows_appended += 1
                
                # Add formula to Clasification column for all newly appended rows (only for Labor_Input tab)
                if tab_name == "Labor_Input" and rows_appended > 0:
                    clasification_col_idx = None
                    for col_idx, header in enumerate(excel_headers, start=1):
                        if header.lower() in ['clasification', 'classification']:
                            clasification_col_idx = col_idx
                            break
                    
                    if clasification_col_idx:
                        # Find Job Title column (column C, which is column index 3)
                        job_title_col_idx = None
                        for col_idx, header in enumerate(excel_headers, start=1):
                            if header.lower() in ['job title', 'job_title']:
                                job_title_col_idx = col_idx
                                break
                        
                        if job_title_col_idx:
                            # Add formula to each new row
                            for i in range(rows_appended):
                                row_num = start_row + i
                                # Formula: =IFERROR(VLOOKUP(C{row_num}, Job_Classification_Lookup!$A$2:$B$100, 2, FALSE), "Other")
                                # Where C{row_num} is the Job Title column at current row
                                job_title_col_letter = get_column_letter(job_title_col_idx)
                                formula = f'=IFERROR(VLOOKUP({job_title_col_letter}{row_num}, Job_Classification_Lookup!$A$2:$B$100, 2, FALSE), "Other")'
                                clasification_cell = worksheet.cell(row=row_num, column=clasification_col_idx)
                                clasification_cell.value = formula
                            
                            print(f"      {CHECKMARK} Added formula to Clasification column for {rows_appended} row(s)")
                
                # Return True for all tabs after successfully appending rows
                if rows_appended > 0:
                    return True
                else:
                    print(f"  {WARNING} No rows to append from CSV file")
                    return False
                    
            except Exception as e:
                print(f"  {CROSS} Error appending CSV to tab '{tab_name}': {e}")
                import traceback
                traceback.print_exc()
                return False
        else:
            # Google Sheets version
            try:
                # Check if tab exists
                if not self.sheet:
                    print(f"  {CROSS} Google Sheet not loaded")
                    return False
                
                try:
                    worksheet = self._get_worksheet(tab_name)
                except gspread.exceptions.WorksheetNotFound:
                    print(f"  {CROSS} Tab '{tab_name}' does not exist in Google Sheet")
                    try:
                        sheet_titles = [ws.title for ws in self._retry_gspread(self.sheet.worksheets)]
                        print(f"     Available tabs: {', '.join(sheet_titles)}")
                    except:
                        pass
                    return False
                
                # Read CSV file
                df = self._read_csv_safe(csv_file)
                if df is None or df.empty:
                    return False
                df = self._filter_total_rows(df, csv_file)
                
                # Get Google Sheets headers (row 1)
                sheet_headers = self._get_sheet_headers(tab_name, worksheet)
                if not sheet_headers:
                    print(f"  {CROSS} No headers found in tab '{tab_name}'")
                    return False
                
                # Date should be the first column
                if sheet_headers[0].lower() not in expected_headers:
                    print(f"  {WARNING} First column should be 'Date', found: {sheet_headers[0]}")
                    # Continue anyway
                
                # Map CSV column names to sheet column names (case-insensitive matching)
                csv_to_sheet_mapping = {}
                unmapped_csv_cols = []
                
                for csv_col in df.columns:
                    csv_col_stripped = str(csv_col).strip()
                    # Skip Clasification column - it will use a formula
                    if csv_col_stripped.lower() in ['clasification', 'classification']:
                        continue
                    
                    # Try exact match first
                    if csv_col_stripped in sheet_headers:
                        csv_to_sheet_mapping[csv_col] = csv_col_stripped
                    else:
                        # Try case-insensitive match
                        found = False
                        for sheet_col in sheet_headers:
                            # Skip Clasification column
                            if sheet_col.lower() in ['clasification', 'classification']:
                                continue
                            if csv_col_stripped.lower() == sheet_col.lower():
                                csv_to_sheet_mapping[csv_col] = sheet_col
                                found = True
                                break
                        if not found:
                            unmapped_csv_cols.append(csv_col)
                
                # Warn about unmapped columns
                if unmapped_csv_cols:
                    print(f"      {WARNING} CSV columns not found in sheet (skipped): {', '.join(unmapped_csv_cols[:5])}{'...' if len(unmapped_csv_cols) > 5 else ''}")
                
                # Add Date column to dataframe (first column)
                date_only = week_ending_date.date() if hasattr(week_ending_date, 'date') else week_ending_date
                df.insert(0, self.date_column_name, date_only)
                csv_to_sheet_mapping[self.date_column_name] = sheet_headers[0]  # Map to first sheet column
                
                # Determine starting row (after existing data)
                all_values = self._retry_gspread(worksheet.get_all_values)
                if not all_values or len(all_values) <= 1:
                    start_row = 2  # Row 1 is header, data starts at row 2
                else:
                    start_row = len(all_values) + 1
                
                # Prepare data rows for batch update
                rows_to_append = []
                for idx, csv_row in df.iterrows():
                    row_data = []
                    # Build row data matching sheet column order
                    for sheet_col in sheet_headers:
                        if sheet_col in csv_to_sheet_mapping.values():
                            # Find the CSV column that maps to this sheet column
                            csv_col = None
                            for csv_key, sheet_val in csv_to_sheet_mapping.items():
                                if sheet_val == sheet_col:
                                    csv_col = csv_key
                                    break
                            
                            if csv_col:
                                if csv_col == self.date_column_name:
                                    value = week_ending_date.date() if hasattr(week_ending_date, 'date') else week_ending_date
                                else:
                                    value = csv_row[csv_col]
                                
                                # Convert value to appropriate format
                                if pd.isna(value):
                                    row_data.append("")
                                elif csv_col == self.date_column_name:
                                    from datetime import date as date_type
                                    if isinstance(value, datetime):
                                        row_data.append(value.date().strftime("%Y-%m-%d"))
                                    elif isinstance(value, date_type):
                                        row_data.append(value.strftime("%Y-%m-%d"))
                                    else:
                                        try:
                                            row_data.append(pd.to_datetime(value).date().strftime("%Y-%m-%d"))
                                        except:
                                            row_data.append(str(value))
                                elif isinstance(value, (int, float)):
                                    row_data.append(float(value))
                                else:
                                    row_data.append(str(value))
                            else:
                                row_data.append("")  # Empty cell for unmapped columns
                        else:
                            row_data.append("")  # Empty cell for columns not in mapping
                    rows_to_append.append(row_data)
                
                # Batch append rows to Google Sheets
                if rows_to_append:
                    self._retry_gspread(worksheet.append_rows, rows_to_append, value_input_option='USER_ENTERED')
                    rows_appended = len(rows_to_append)
                    
                    # Add formula to Clasification column for all newly appended rows (only for Labor_Input tab)
                    if tab_name == "Labor_Input" and rows_appended > 0:
                        clasification_col_idx = None
                        for col_idx, header in enumerate(sheet_headers, start=1):
                            if header.lower() in ['clasification', 'classification']:
                                clasification_col_idx = col_idx
                                break
                        
                        if clasification_col_idx:
                            # Find Job Title column
                            job_title_col_idx = None
                            for col_idx, header in enumerate(sheet_headers, start=1):
                                if header.lower() in ['job title', 'job_title']:
                                    job_title_col_idx = col_idx
                                    break
                            
                            if job_title_col_idx:
                                # Add formula to each new row
                                job_title_col_letter = self._column_index_to_a1(job_title_col_idx)
                                clasification_col_letter = self._column_index_to_a1(clasification_col_idx)
                                
                                for i in range(rows_appended):
                                    row_num = start_row + i
                                    # Formula: =IFERROR(VLOOKUP(C{row_num}, Job_Classification_Lookup!$A$2:$B$100, 2, FALSE), "Other")
                                    formula = f'=IFERROR(VLOOKUP({job_title_col_letter}{row_num}, Job_Classification_Lookup!$A$2:$B$100, 2, FALSE), "Other")'
                                    cell_range = f"{clasification_col_letter}{row_num}"
                                    self._retry_gspread(
                                        worksheet.update,
                                        range_name=cell_range,
                                        values=[[formula]],
                                        value_input_option='USER_ENTERED'
                                    )
                                
                                print(f"      {CHECKMARK} Added formula to Clasification column for {rows_appended} row(s)")
                    
                    # Success message is handled by calling function
                    return True
                else:
                    print(f"  {WARNING} No data to append")
                    return False
                
            except Exception as e:
                print(f"  {CROSS} Error appending CSV to tab '{tab_name}': {e}")
                import traceback
                traceback.print_exc()
                return False
    
    def validate_configuration(self, verbose: bool = True) -> bool:
        """Validate configuration without connecting to Google Sheets."""
        if verbose:
            print("Validating configuration...")
        
        # Check test mode
        if verbose:
            if self.test_mode:
                print(f"  {CHECKMARK} TEST MODE enabled - using local Excel file")
            else:
                print(f"  {CHECKMARK} PRODUCTION MODE - using Google Sheets")
        
        # Check config structure
        required_fields = ['google_sheet', 'csv_folder']
        for field in required_fields:
            if field not in self.config:
                if verbose:
                    print(f"  {CROSS} Missing required field: {field}")
                return False
            if verbose:
                print(f"  {CHECKMARK} {field} configured")
        
        if self.test_mode:
            # Validate Excel file configuration
            excel_path = Path(__file__).parent / self.excel_file
            
            # Search for Excel file in common locations
            if not excel_path.exists():
                base_path = Path(__file__).parent
                found_path = None
                
                # Check subdirectories (exclude backup files)
                for subdir in base_path.iterdir():
                    if subdir.is_dir():
                        potential_path = subdir / self.excel_file
                        # Skip backup files
                        if potential_path.exists() and "backup" not in potential_path.name.lower():
                            found_path = potential_path
                            break
                
                if found_path:
                    excel_path = found_path
                    print(f"  {CHECKMARK} Excel file found: {excel_path.name}")
                    print(f"     Location: {excel_path}")
                else:
                    print(f"  {WARNING} Excel file not found: {self.excel_file}")
                    print(f"     Searched in: {base_path} and subdirectories")
                    print(f"     Note: Backup files are excluded from search")
                    print(f"     The file will be searched again when loading")
            else:
                # Double-check: Don't use backup files
                if "backup" in excel_path.name.lower():
                    print(f"  {WARNING} Excel file is a backup file: {excel_path.name}")
                    print(f"     Searching for original file in subdirectories...")
                else:
                    print(f"  {CHECKMARK} Excel file found: {excel_path.name}")
        else:
            # Check Google Sheet config (only if not in test mode)
            if 'sheet_id' not in self.config['google_sheet']:
                print(f"  {CROSS} Missing google_sheet.sheet_id")
                return False
            
            sheet_id = self.config['google_sheet']['sheet_id']
            if sheet_id == "YOUR_GOOGLE_SHEET_ID_HERE":
                print(f"  {CROSS} Google Sheet ID not configured (still using placeholder)")
                return False
            
            print(f"  {CHECKMARK} Google Sheet ID configured: {sheet_id[:20]}...")
            
            # Check credentials based on auth method
            auth_method = self.config.get('auth_method', 'service_account').lower()
            if auth_method == 'oauth':
                # Check if OAuth credentials are in secrets.json
                oauth_credentials = self.config.get('_oauth_credentials')
                if oauth_credentials:
                    print(f"  {CHECKMARK} OAuth credentials found in secrets.json")
                else:
                    # Fall back to checking file
                    oauth_creds_path = self.config.get('oauth_credentials_file', 'oauth_credentials.json')
                    if os.path.exists(oauth_creds_path):
                        print(f"  {CHECKMARK} OAuth credentials file found: {oauth_creds_path}")
                    else:
                        print(f"  {CROSS} OAuth credentials not found!")
                        print(f"     Add 'oauth_credentials' to secrets.json or create: {oauth_creds_path}")
                        if not self.dry_run:
                            return False
            else:
                credentials_path = self.config.get('credentials_file', 'credentials.json')
                if os.path.exists(credentials_path):
                    print(f"  {CHECKMARK} Service account credentials file found: {credentials_path}")
                else:
                    print(f"  {CROSS} Service account credentials file not found: {credentials_path}")
                    if not self.dry_run:
                        return False
        
        # Check CSV folder (test mode uses local folder; production uses Drive download)
        if self.test_mode:
            csv_folder = self.get_csv_folder_path()
            if not csv_folder.exists():
                if verbose:
                    print(f"  {WARNING} CSV folder does not exist: {csv_folder}")
                    print(f"     It will be created when you run the script")
            else:
                if verbose:
                    print(f"  {CHECKMARK} CSV folder found: {csv_folder.name}")
        else:
            if verbose:
                print(f"  {CHECKMARK} Sales input will be pulled from Drive folder: {self.drive_sales_input_id}")
                print(f"  {CHECKMARK} Labor input will be pulled from Drive folder: {self.drive_labor_input_id}")
        
        if verbose:
            print(f"\n  {CHECKMARK} Configuration validation complete!")
        return True
    
    def process_csv_files(self, skip_validation: bool = False) -> None:
        """Main processing function - finds and processes all CSV files."""
        if not skip_validation:
            if self.dry_run:
                print("\n" + "="*70)
                print("VALIDATION")
                print("="*70)
                if not self.validate_configuration():
                    print(f"\n{WARNING} Configuration validation failed. Please fix errors above.")
                    return
            else:
                print("\n" + "="*70)
                print("VALIDATION")
                print("="*70)
                if not self.validate_configuration():
                    print(f"\n{WARNING} Configuration validation failed. Please fix errors above.")
                    return
                
                # Load appropriate target based on test mode
                if self.test_mode:
                    if not self.load_excel_file():
                        print(f"\n{WARNING} Failed to load Excel file. Exiting.")
                        return
                else:
                    if not self.authenticate_google_sheets():
                        return
        elif not self.test_mode or not self.workbook:
            # If skipping validation but Excel not loaded, load it
            if self.test_mode and not self.workbook:
                if not self.load_excel_file():
                    print(f"\n{WARNING} Failed to load Excel file. Exiting.")
                    return
        
        print("\n" + "="*70)
        print("PROCESSING SALES CSV FILES")
        print("="*70)
        
        # Determine which folders/dates to process
        if self.test_mode:
            if self.process_oldest:
                missing_folders = self.find_missing_sales_folders()
                if not missing_folders:
                    print(f"\n{CROSS} Could not find oldest missing week.")
                    print(f"  All available weeks may already exist in the Excel file.")
                    return
                folders_to_process = missing_folders
                print(f"\n  Found {len(folders_to_process)} missing date(s) to process (oldest first)")
            else:
                missing_folders = self.find_missing_sales_folders()
                if missing_folders:
                    folders_to_process = missing_folders
                    print(f"\n  Found {len(folders_to_process)} missing date(s) to process")
                else:
                    # Fall back to latest folder
                    csv_folder = self.get_csv_folder_path()
                    input_date = self.extract_week_ending_date()
                    if not input_date:
                        print(f"\n{CROSS} Could not extract input date from folder name.")
                        print(f"  Expected format: SalesSummary_YYYY-MM-DD_YYYY-MM-DD (start date - end date)")
                        return
                    folders_to_process = [(csv_folder, input_date)]
        else:
            drive_folders = self.find_all_sales_drive_folders_with_dates()
            if self.process_oldest:
                missing_folders = self.find_missing_sales_drive_folders()
                if not missing_folders:
                    print(f"\n{CROSS} Could not find oldest missing week.")
                    print(f"  All available weeks may already exist in the Google Sheet.")
                    return
                folders_to_process = missing_folders
                print(f"\n  Found {len(folders_to_process)} missing date(s) to process (oldest first)")
            else:
                missing_folders = self.find_missing_sales_drive_folders()
                if missing_folders:
                    folders_to_process = missing_folders
                    print(f"\n  Found {len(folders_to_process)} missing date(s) to process")
                else:
                    if not drive_folders:
                        print(f"\n{CROSS} No SalesSummary folders found in Drive Sales_Input folder.")
                        return
                    folders_to_process = [drive_folders[-1]]

        # Process each folder/date
        for folder_idx, folder_entry in enumerate(folders_to_process, 1):
            if self.test_mode:
                folder_path, input_date = folder_entry
            else:
                folder_id, folder_name, input_date = folder_entry
                folder_path = self._prepare_sales_input_folder_from_drive(folder_id, folder_name)

            if not folder_path.exists():
                print(f"\n  {WARNING} Folder does not exist: {folder_path}")
                continue

            csv_files = self.find_csv_files(folder_path)
            if not csv_files:
                continue

            input_date_str = input_date.strftime("%Y-%m-%d")
            print(f"\n  [{folder_idx}/{len(folders_to_process)}] Date: {input_date_str}")

            # Filter CSV files to only process the 4 mapped files
            files_to_process = []
            for csv_file in csv_files:
                csv_filename = csv_file.name
                if csv_filename in self.csv_to_tab_mapping:
                    files_to_process.append(csv_file)

            if not files_to_process:
                print(f"\n  {CROSS} No CSV files found matching the required files:")
                for csv_name in self.csv_to_tab_mapping.keys():
                    print(f"      - {csv_name}")
                continue

            print(f"\n  Files to process: {len(files_to_process)} CSV file(s)\n")

            # Process each CSV file
            for idx, csv_file in enumerate(files_to_process, 1):
                csv_filename = csv_file.name
                tab_name = self.csv_to_tab_mapping.get(csv_filename)
            
                if not tab_name:
                    print(f"\n  {CROSS} No tab mapping found for: {csv_filename}")
                    continue
            
                print(f"  [{idx}/{len(files_to_process)}] {csv_filename}")
                print(f"      -> Tab: {tab_name}")
                
                # Check if tab exists
                if self.test_mode:
                    if not self.workbook or tab_name not in self.workbook.sheetnames:
                        print(f"  {CROSS} Tab '{tab_name}' does not exist in Excel file")
                        print(f"     Available tabs: {', '.join(self.workbook.sheetnames) if self.workbook else 'N/A'}")
                        continue
                else:
                    if not self.sheet:
                        print(f"  {CROSS} Google Sheet not loaded")
                        continue
                    try:
                        self._get_worksheet(tab_name)  # Check if worksheet exists
                    except gspread.exceptions.WorksheetNotFound:
                        try:
                            sheet_titles = [ws.title for ws in self.sheet.worksheets()]
                            print(f"  {CROSS} Tab '{tab_name}' does not exist in Google Sheet")
                            print(f"     Available tabs: {', '.join(sheet_titles)}")
                        except:
                            print(f"  {CROSS} Tab '{tab_name}' does not exist in Google Sheet")
                        continue
                
                # Check if date already exists
                exists, row_count = self.check_week_ending_exists(tab_name, input_date)
                
                if exists:
                    # Ask user for override confirmation
                    if self.dry_run:
                        print(f"  [DRY RUN] Would ask user to override {row_count} existing row(s)")
                        continue
                    
                    should_override = self.ask_user_override(tab_name, input_date, row_count)
                    
                    if not should_override:
                        print(f"      [SKIP] Skipped (user chose not to override)\n")
                        continue
                    
                    # Delete existing rows
                    deleted_count = self.delete_rows_with_week_ending(tab_name, input_date)
                    print(f"      {CHECKMARK} Deleted {deleted_count} existing row(s)")
                
                # Append CSV data to tab
                if self.dry_run:
                    print(f"      [DRY RUN] Would append CSV data\n")
                else:
                    success = self.append_csv_to_excel_tab(csv_file, tab_name, input_date)
                    
                    if success:
                        if self.test_mode:
                            # Save workbook
                            try:
                                self.workbook.save(self.excel_file_path)
                                print(f"      {CHECKMARK} Completed and saved\n")
                            except PermissionError:
                                print(f"      {WARNING} Could not save Excel file (file may be open)")
                                print(f"      {WARNING} Please close the Excel file and run the script again to save changes\n")
                        else:
                            # Google Sheets - changes are saved automatically
                            print(f"      {CHECKMARK} Completed\n")
                    else:
                        print(f"      {CROSS} Failed to append CSV data\n")
    
    def ask_user_which_file_to_process(self, latest_file: Path, duplicate_files: List[Path], week_ending_date: datetime) -> Optional[Path]:
        """Ask user which file to process when multiple files exist for the same input date."""
        week_ending_str = week_ending_date.strftime("%Y-%m-%d")
        print(f"\n  {WARNING} Multiple CSV files found for date {week_ending_str}:")
        print(f"    1. {latest_file.name} (latest)")
        
        for idx, dup_file in enumerate(duplicate_files, start=2):
            print(f"    {idx}. {dup_file.name}")
        
        while True:
            try:
                choice = input(f"  Which file do you want to process? (1-{len(duplicate_files) + 1}, or 'skip'): ").strip().lower()
                
                if choice == 'skip':
                    return None
                
                choice_num = int(choice)
                if choice_num == 1:
                    return latest_file
                elif 2 <= choice_num <= len(duplicate_files) + 1:
                    return duplicate_files[choice_num - 2]
                else:
                    print(f"  Please enter a number between 1 and {len(duplicate_files) + 1}, or 'skip'")
            except ValueError:
                print(f"  Please enter a valid number or 'skip'")
    
    def process_labor_input_csv_files(self, skip_validation: bool = False) -> None:
        """Process PayrollExport CSV files from Labor_Input folder."""
        if not skip_validation:
            if self.dry_run:
                print("\n" + "="*70)
                print("VALIDATION")
                print("="*70)
                if not self.validate_configuration():
                    print(f"\n{WARNING} Configuration validation failed. Please fix errors above.")
                    return
            else:
                print("\n" + "="*70)
                print("VALIDATION")
                print("="*70)
                if not self.validate_configuration():
                    print(f"\n{WARNING} Configuration validation failed. Please fix errors above.")
                    return
            
            # Load appropriate file based on test mode
            if self.test_mode:
                # Load Excel file (even in dry run mode to validate tab exists)
                if not self.load_excel_file():
                    print(f"\n{WARNING} Failed to load Excel file. Exiting.")
                    return
            else:
                # Load Google Sheet
                if not self.authenticate_google_sheets():
                    print(f"\n{WARNING} Failed to authenticate Google Sheets. Exiting.")
                    return
        elif not self.test_mode or not self.workbook:
            # If skipping validation but Excel not loaded, load it
            if self.test_mode and not self.workbook:
                if not self.load_excel_file():
                    print(f"\n{WARNING} Failed to load Excel file. Exiting.")
                    return
        
        # Get Labor_Input folder path
        if self.test_mode:
            base_path = Path(__file__).parent
            labor_input_folder = base_path / "Labor_Input"
        else:
            labor_input_folder = self._prepare_labor_input_folder_from_drive()
        
        print("\n" + "="*70)
        print("PROCESSING LABOR INPUT CSV FILES")
        print("="*70)
        
        # Determine which labor files to process
        if self.process_oldest:
            # Find all missing dates (oldest first)
            missing_files = self.find_missing_labor_csvs(labor_input_folder)
            if not missing_files:
                target_name = "Excel file" if self.test_mode else "Google Sheet"
                print(f"\n{CROSS} Could not find oldest missing PayrollExport CSV file.")
                print(f"  All available weeks may already exist in the {target_name}.")
                return
            files_to_process = missing_files
            print(f"\n  Found {len(files_to_process)} missing date(s) to process (oldest first)")
        else:
            missing_files = self.find_missing_labor_csvs(labor_input_folder)
            if missing_files:
                files_to_process = missing_files
                print(f"\n  Found {len(files_to_process)} missing date(s) to process")
            else:
                # Find latest CSV file (fallback behavior)
                latest_file, week_ending_date, duplicate_files = self.find_latest_labor_input_csv(labor_input_folder)
                
                if not latest_file or not week_ending_date:
                    print(f"\n{CROSS} Could not find or process latest PayrollExport CSV file.")
                    return
                
                # Check for duplicate files with same input date
                if duplicate_files:
                    if self.dry_run:
                        print(f"\n  [DRY RUN] Would ask user to choose from {len(duplicate_files) + 1} file(s)")
                        return
                    
                    file_to_process = self.ask_user_which_file_to_process(latest_file, duplicate_files, week_ending_date)
                    
                    if not file_to_process:
                        print(f"\n  [SKIP] Skipped (user chose to skip)\n")
                        return
                else:
                    file_to_process = latest_file
                
                files_to_process = [(file_to_process, week_ending_date, [])]
        
        # Check if Labor_Input tab exists
        tab_name = "Labor_Input"
        if self.test_mode:
            if not self.workbook or tab_name not in self.workbook.sheetnames:
                print(f"\n  {CROSS} Tab '{tab_name}' does not exist in Excel file")
                if self.workbook:
                    print(f"     Available tabs: {', '.join(self.workbook.sheetnames)}")
                return
        else:
            if not self.sheet:
                print(f"\n  {CROSS} Google Sheet not loaded")
                return
            try:
                self._get_worksheet(tab_name)  # Check if worksheet exists
                self._get_worksheet(tab_name)  # Check if worksheet exists
            except gspread.exceptions.WorksheetNotFound:
                try:
                    sheet_titles = [ws.title for ws in self.sheet.worksheets()]
                    print(f"\n  {CROSS} Tab '{tab_name}' does not exist in Google Sheet")
                    print(f"     Available tabs: {', '.join(sheet_titles)}")
                except:
                    print(f"\n  {CROSS} Tab '{tab_name}' does not exist in Google Sheet")
                return
        
        for idx, (file_to_process, week_ending_date, duplicate_files) in enumerate(files_to_process, 1):
            week_ending_str = week_ending_date.strftime("%Y-%m-%d")
            print(f"\n  [{idx}/{len(files_to_process)}] CSV File: {file_to_process.name}")
            print(f"  Date: {week_ending_str}")
            
            if duplicate_files:
                dup_names = ", ".join([f.name for f in duplicate_files[:5]])
                print(f"  {WARNING} Multiple files found for {week_ending_str}; using latest. Duplicates: {dup_names}{'...' if len(duplicate_files) > 5 else ''}")
            
            # Check if date already exists
            exists, row_count = self.check_week_ending_exists(tab_name, week_ending_date)
            
            if exists:
                if self.dry_run:
                    print(f"\n  [DRY RUN] Would ask user to override {row_count} existing row(s)")
                    print(f"  [DRY RUN] Would append CSV data to tab '{tab_name}'\n")
                    continue
                
                # Ask user for override confirmation
                should_override = self.ask_user_override(tab_name, week_ending_date, row_count)
                
                if not should_override:
                    print(f"\n  [SKIP] Skipped (user chose not to override)\n")
                    continue
                
                # Delete existing rows
                deleted_count = self.delete_rows_with_week_ending(tab_name, week_ending_date)
                print(f"      {CHECKMARK} Deleted {deleted_count} existing row(s)")
            
            # Append CSV data to tab
            print(f"\n  Processing: {file_to_process.name}")
            print(f"      -> Tab: {tab_name}")
            
            if self.dry_run:
                print(f"      [DRY RUN] Would append CSV data\n")
            else:
                success = self.append_csv_to_excel_tab(file_to_process, tab_name, week_ending_date)
                
                if success:
                    if self.test_mode:
                        # Save workbook
                        try:
                            self.workbook.save(self.excel_file_path)
                            print(f"      {CHECKMARK} Completed and saved\n")
                        except PermissionError:
                            print(f"      {WARNING} Could not save Excel file (file may be open)")
                            print(f"      {WARNING} Please close the Excel file and run the script again to save changes\n")
                    else:
                        # Google Sheets - changes are saved automatically
                        print(f"      {CHECKMARK} Completed\n")
                else:
                    print(f"      {CROSS} Failed to append CSV data\n")
    
    def process_all_csv_files(self) -> None:
        """Process both Sales Input CSV files and Labor_Input CSV files."""
        print("\n" + "="*70)
        print("PROCESSING ALL CSV FILES")
        print("="*70)
        print("  • Sales Input CSV Files")
        print("  • Labor Input CSV Files")
        
        # Validate once at the beginning
        if self.dry_run:
            print("\n" + "="*70)
            print("VALIDATION")
            print("="*70)
            if not self.validate_configuration():
                print(f"\n{WARNING} Configuration validation failed. Please fix errors above.")
                return
        else:
            print("\n" + "="*70)
            print("VALIDATION")
            print("="*70)
            if not self.validate_configuration():
                print(f"\n{WARNING} Configuration validation failed. Please fix errors above.")
                return
            
            # Load Excel file if in test mode
            if self.test_mode:
                if not self.load_excel_file():
                    print(f"\n{WARNING} Failed to load Excel file. Exiting.")
                    return
            else:
                if not self.authenticate_google_sheets():
                    return
        
        # Process Sales Input CSV files first (skip validation since we already did it)
        print("\n" + "="*70)
        print("STEP 1: SALES INPUT CSV FILES")
        print("="*70)
        self.process_csv_files(skip_validation=True)
        
        # Process Labor Input CSV files (skip validation since we already did it)
        print("\n" + "="*70)
        print("STEP 2: LABOR INPUT CSV FILES")
        print("="*70)
        self.process_labor_input_csv_files(skip_validation=True)
        
        print("\n" + "="*70)
        print("ALL PROCESSING COMPLETE!")
        print("="*70)

def show_interactive_menu() -> Tuple[str, str, str]:
    """Show interactive menu and return user's choices.
    Returns: (process_type, week_type, mode_type)"""
    print("\n" + "="*70)
    print("CSV TO SHEETS AUTOMATION - INTERACTIVE MENU")
    print("="*70)
    print("\nSelect an option:")
    print()
    print("  SALES INPUT:")
    print("    1. Sales + Latest + Testing")
    print("    2. Sales + Oldest + Testing")
    print("    3. Sales + Latest + Production")
    print("    4. Sales + Oldest + Production")
    print()
    print("  LABOR INPUT:")
    print("    5. Labor + Latest + Testing")
    print("    6. Labor + Oldest + Testing")
    print("    7. Labor + Latest + Production")
    print("    8. Labor + Oldest + Production")
    print()
    print("  BOTH (Sales + Labor):")
    print("    9. All + Latest + Testing")
    print("   10. All + Oldest + Testing")
    print("   11. All + Latest + Production")
    print("   12. All + Oldest + Production")
    print()
    
    while True:
        try:
            choice = input("Enter your choice (1-12): ").strip()
            choice_num = int(choice)
            
            if choice_num == 1:
                return ("sales", "latest", "testing")
            elif choice_num == 2:
                return ("sales", "oldest", "testing")
            elif choice_num == 3:
                return ("sales", "latest", "production")
            elif choice_num == 4:
                return ("sales", "oldest", "production")
            elif choice_num == 5:
                return ("labor", "latest", "testing")
            elif choice_num == 6:
                return ("labor", "oldest", "testing")
            elif choice_num == 7:
                return ("labor", "latest", "production")
            elif choice_num == 8:
                return ("labor", "oldest", "production")
            elif choice_num == 9:
                return ("all", "latest", "testing")
            elif choice_num == 10:
                return ("all", "oldest", "testing")
            elif choice_num == 11:
                return ("all", "latest", "production")
            elif choice_num == 12:
                return ("all", "oldest", "production")
            else:
                print("  Please enter a number between 1 and 12")
        except ValueError:
            print("  Please enter a valid number")
        except KeyboardInterrupt:
            print("\n\nExiting...")
            sys.exit(0)

def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(description='CSV to Google Sheets Automation')
    parser.add_argument('--dry-run', action='store_true', 
                       help='Validate configuration and show what would be uploaded without actually uploading')
    parser.add_argument('--config', default='config.json',
                       help='Path to configuration file (default: config.json)')
    
    # Process type flags (mutually exclusive)
    process_group = parser.add_mutually_exclusive_group()
    process_group.add_argument('--sales', action='store_true',
                              help='Process Sales Input CSV files only')
    process_group.add_argument('--labor', action='store_true',
                              help='Process Labor_Input PayrollExport CSV files only')
    process_group.add_argument('--all', action='store_true',
                              help='Process both Sales Input and Labor_Input CSV files')
    # Keep --labor-input as alias for backward compatibility
    process_group.add_argument('--labor-input', action='store_true',
                              help='[DEPRECATED] Use --labor instead. Process Labor_Input PayrollExport CSV files only')
    
    # Week selection flags (mutually exclusive)
    week_group = parser.add_mutually_exclusive_group()
    week_group.add_argument('--latest', action='store_true',
                           help='Process the latest week (default)')
    week_group.add_argument('--oldest', action='store_true',
                           help='Process the oldest missing week')
    
    # Mode selection flags (mutually exclusive)
    mode_group = parser.add_mutually_exclusive_group()
    mode_group.add_argument('--testing', action='store_true',
                           help='Use testing mode (Excel file)')
    mode_group.add_argument('--prod', action='store_true',
                           help='Use production mode (Google Sheets)')
    
    args = parser.parse_args()
    
    # Handle deprecated --labor-input flag
    if args.labor_input:
        args.labor = True
    
    # Determine process type, week type, and mode type
    process_type = None
    week_type = "latest"  # default
    mode_type = None  # will use config.json if not specified
    
    # Process type
    if args.sales:
        process_type = "sales"
    elif args.labor:
        process_type = "labor"
    elif args.all:
        process_type = "all"
    
    # Week type
    if args.oldest:
        week_type = "oldest"
    elif args.latest:
        week_type = "latest"
    # else: default to "latest"
    
    # Mode type
    if args.testing:
        mode_type = True  # testing mode
    elif args.prod:
        mode_type = False  # production mode
    # else: None (use config.json)
    
    # If no process type specified, show interactive menu
    if process_type is None:
        process_type, week_type, mode_type_str = show_interactive_menu()
        mode_type = True if mode_type_str == "testing" else False
    
    print("="*60)
    print("CSV to Google Sheets Automation")
    print("="*60)
    
    # Convert week_type to boolean
    process_oldest = (week_type == "oldest")
    
    try:
        automation = CSVToSheetsAutomation(
            config_path=args.config, 
            dry_run=args.dry_run, 
            process_oldest=process_oldest,
            mode_override=mode_type
        )
        
        if process_type == "all":
            automation.process_all_csv_files()
        elif process_type == "labor":
            automation.process_labor_input_csv_files()
        else:  # sales (default)
            automation.process_csv_files()
    except FileNotFoundError as e:
        print(f"\nError: {e}")
        print("Please create the configuration file. See README.md for details.")
        sys.exit(1)
    except Exception as e:
        print(f"\nUnexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
